import io
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from core.pagination import StandardPagination
from core.permissions import DataScopeMixin, validate_branches_in_scope
from apps.permissions.permissions import OperationPermission
from apps.audit.decorators import audit_create
from rest_framework.exceptions import ValidationError
from .models import AssetStock, FixedAsset, LedgerAdjustment
from .serializers import (
    AssetStockSerializer,
    FixedAssetSerializer,
    FixedAssetSupplementSerializer,
    LedgerAdjustmentSerializer,
)
from .filters import AssetStockFilterSet, FixedAssetFilterSet


class AssetStockViewSet(DataScopeMixin, viewsets.ModelViewSet):
    """资产汇总台账 —— 库存唯一事实源（只读 + 增量导入）。

    行级写接口（新增/编辑/删除/批量删除）已下线：数量变动唯一通道是流转单与
    调整单（铁律 2）。导入为两段式增量：默认出差异预览，confirm=1 时按差异
    逐行生成调整单（事由=导入调整）入账。
    """

    queryset = AssetStock.objects.select_related('branch', 'item').all()
    serializer_class = AssetStockSerializer
    filterset_class = AssetStockFilterSet
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    scope_branch_field = 'branch'
    # 导入权限在 import_excel 内校验（manage_assets 或 adjust_ledger 任一）

    # 台账增量导入模板 3 列：分公司 / 资产编号 / 在库数量
    STOCK_TEMPLATE_HEADERS = ['分公司', '资产编号', '在库数量']

    def get_queryset(self):
        qs = super().get_queryset()
        return self.get_scoped_queryset(qs)

    # ---- 行级写接口下线（铁律 2：数量变动须经单据） ----

    def _write_disabled(self, request):
        return Response(
            {'detail': '台账为库存唯一事实源，数量变动须经流转单或调整单'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def create(self, request, *args, **kwargs):
        return self._write_disabled(request)

    def update(self, request, *args, **kwargs):
        return self._write_disabled(request)

    def partial_update(self, request, *args, **kwargs):
        return self._write_disabled(request)

    def destroy(self, request, *args, **kwargs):
        return self._write_disabled(request)

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        return self._write_disabled(request)

    # ---- 模板 / 导入 / 导出 ----

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """下载台账增量导入模板（分公司 / 资产编号 / 在库数量）。"""
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产汇总'
        ws.append(self.STOCK_TEMPLATE_HEADERS)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="asset_summary_template.xlsx"'
        return response

    def _parse_import_rows(self, file, user):
        """解析增量导入文件 → (diffs, errors)。diffs 元素含 branch/item ORM 对象。"""
        from apps.categories.models import Category
        from apps.categories.views import suggest_similar_codes
        from apps.organizations.models import Branch

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            raise ValueError(f'文件解析失败: {str(e)}')

        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []

        header_row = [str(c or '').strip() for c in all_rows[0]]
        col = {}
        for idx, name in enumerate(header_row):
            if name and name not in col:
                col[name] = idx

        def cell(row, name):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return ''
            val = row[idx]
            return '' if val is None else val

        diffs = []
        errors = []
        seen_keys = set()
        branch_map = {b.name: b for b in Branch.objects.all()}
        item_map = {c.asset_code: c for c in Category.objects.all()}

        for i, row in enumerate(all_rows[1:], start=2):
            asset_code = str(cell(row, '资产编号')).strip()
            branch_name = str(cell(row, '分公司')).strip()
            if not asset_code and not branch_name:
                continue
            if not asset_code:
                errors.append(f'第 {i} 行: 资产编号为空')
                continue
            item = item_map.get(asset_code)
            if item is None:
                similar = suggest_similar_codes(asset_code)
                hint = f'，是否想找：{"、".join(similar)}' if similar else ''
                errors.append(f'第 {i} 行: 资产编号 {asset_code} 未在品目字典登记{hint}')
                continue
            branch = branch_map.get(branch_name)
            if branch is None:
                errors.append(f'第 {i} 行: 分公司「{branch_name}」不存在')
                continue
            try:
                validate_branches_in_scope(user, branch)
            except ValidationError:
                errors.append(f'第 {i} 行: 分公司「{branch_name}」不在你的授权范围')
                continue
            key = (branch_name, asset_code)
            if key in seen_keys:
                errors.append(f'第 {i} 行: 资产编号 {asset_code} 在文件内重复')
                continue
            seen_keys.add(key)

            qty_raw = cell(row, '在库数量')
            try:
                target_qty = int(qty_raw)
            except (ValueError, TypeError):
                errors.append(f'第 {i} 行: 在库数量 "{qty_raw}" 不是有效整数')
                continue

            current = AssetStock.objects.filter(branch=branch, item=item).first()
            current_qty = current.在库数量 if current else 0
            delta = target_qty - current_qty
            if delta != 0:
                diffs.append({
                    'row': i,
                    'branch': branch,
                    'branch_name': branch_name,
                    'item': item,
                    '资产编号': asset_code,
                    '资产名称': item.asset_name,
                    '现值': current_qty,
                    '导入值': target_qty,
                    '变动量': delta,
                })
        return diffs, errors

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='import')
    def import_excel(self, request):
        """台账增量导入：默认返回差异预览；confirm=1 时逐差异生成调整单。"""
        from core.upload_validation import validate_excel_upload, UploadValidationError

        if not (request.user.can('adjust_ledger') or request.user.can('manage_assets')):
            return Response(
                {'detail': '台账导入需要台账调整或资产管理权限'},
                status=status.HTTP_403_FORBIDDEN,
            )

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            validate_excel_upload(file)
        except UploadValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            diffs, errors = self._parse_import_rows(file, request.user)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if str(request.data.get('confirm')) in ('1', 'true', 'True'):
            from .services import ledger
            applied = 0
            for d in diffs:
                ledger.apply_adjustment(
                    branch=d['branch'],
                    item=d['item'],
                    column=ledger.COLUMN_STOCK,
                    delta=d['变动量'],
                    reason=f'导入调整（在库 {d["现值"]} → {d["导入值"]}）',
                    operator=request.user,
                )
                applied += 1
            return Response({'applied': applied, 'errors': errors})

        return Response({
            'diffs': [
                {
                    '行号': d['row'],
                    '分公司': d['branch_name'],
                    '资产编号': d['资产编号'],
                    '资产名称': d['资产名称'],
                    '现值': d['现值'],
                    '导入值': d['导入值'],
                    '变动量': d['变动量'],
                }
                for d in diffs
            ],
            'errors': errors,
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """台账导出（列同页面表头，序号为行号）。"""
        import openpyxl
        from django.http import HttpResponse

        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产汇总'

        headers = [
            '序号', '分公司', '资产编号', '资产名称', '规格', '资产类目', '物品分类',
            '管理方式', '在库数量', '在用数量', '回收库数量', '总量', '警戒线', '是否充足',
        ]
        ws.append(headers)

        for idx, stock in enumerate(queryset, start=1):
            ws.append([
                idx,
                stock.branch.name,
                stock.item.asset_code,
                stock.item.asset_name,
                stock.item.specification,
                stock.item.asset_category,
                stock.item.item_category,
                stock.item.get_management_type_display(),
                stock.在库数量,
                stock.在用数量,
                stock.回收库数量,
                stock.总量,
                stock.生效警戒线 if stock.生效警戒线 is not None else '',
                '是' if stock.是否充足 else '否',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="asset_summary.xlsx"'
        return response


class LedgerAdjustmentViewSet(DataScopeMixin, viewsets.ReadOnlyModelViewSet):
    """台账调整单：创建即生效（走 ledger service 唯一写入口），列表按范围可读。"""

    queryset = LedgerAdjustment.objects.select_related(
        'branch', 'item', '经办人', 'source_task',
    ).all()
    serializer_class = LedgerAdjustmentSerializer
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    scope_branch_field = 'branch'
    required_operations = {
        'create': 'adjust_ledger',
    }

    def get_queryset(self):
        qs = super().get_queryset()
        qs = self.get_scoped_queryset(qs)
        params = self.request.query_params
        branch = params.get('branch')
        if branch:
            qs = qs.filter(branch_id=branch)
        asset_code = (params.get('assetCode') or '').strip()
        if asset_code:
            qs = qs.filter(item__asset_code__icontains=asset_code)
        date_from = params.get('dateFrom')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        date_to = params.get('dateTo')
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    @audit_create(resource_type='LedgerAdjustment')
    def create(self, request, *args, **kwargs):
        from .services import ledger
        from apps.categories.models import Category
        from apps.organizations.models import Branch

        branch_id = request.data.get('branch')
        asset_code = (request.data.get('资产编号') or '').strip()
        column = request.data.get('目标列')
        reason = (request.data.get('事由') or '').strip()
        try:
            delta = int(request.data.get('变动量'))
        except (ValueError, TypeError):
            return Response({'detail': '变动量必须是整数'}, status=status.HTTP_400_BAD_REQUEST)

        branch = Branch.objects.filter(id=branch_id).first() if branch_id else None
        if branch is None:
            branch_name = (request.data.get('分公司') or '').strip()
            branch = Branch.objects.filter(name=branch_name).first() if branch_name else None
        if branch is None:
            return Response({'detail': '分公司无效'}, status=status.HTTP_400_BAD_REQUEST)
        validate_branches_in_scope(request.user, branch)
        item = Category.objects.filter(asset_code=asset_code).first() if asset_code else None
        if item is None:
            return Response(
                {'detail': f'资产编号 {asset_code} 未在品目字典登记'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not reason:
            return Response({'detail': '事由不能为空'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            adjustment = ledger.apply_adjustment(
                branch=branch,
                item=item,
                column=column,
                delta=delta,
                reason=reason,
                operator=request.user,
            )
        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(adjustment).data, status=status.HTTP_201_CREATED)


class FixedAssetViewSet(DataScopeMixin, viewsets.ReadOnlyModelViewSet):
    """固定资产实例视图 —— P2 第二刀重塑：冻结只读 + 序列号补录 + 生平查询。

    实例出生 = 采购单（或存量迁移），状态/使用人/分公司变动经流转单
    （services/instances.py 由台账唯一写入口同事务调用，架构测试执法）。
    """
    queryset = FixedAsset.objects.select_related(
        'branch', 'department', 'item', 'birth_line__transfer',
    ).all()
    serializer_class = FixedAssetSerializer
    filterset_class = FixedAssetFilterSet
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    scope_branch_field = 'branch'
    required_operations = {
        'supplement': 'manage_instances',
    }

    def get_queryset(self):
        qs = super().get_queryset()
        return self.get_scoped_queryset(qs)

    def _frozen(self, request):
        return Response(
            {'detail': '实例变动请经流转单（出生=采购入库单，存量=系统迁移）；'
                       '序列号待补录请使用补录操作'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def create(self, request, *args, **kwargs):
        return self._frozen(request)

    def update(self, request, *args, **kwargs):
        return self._frozen(request)

    def partial_update(self, request, *args, **kwargs):
        return self._frozen(request)

    def destroy(self, request, *args, **kwargs):
        return self._frozen(request)

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        return self._frozen(request)

    # 实例表导出列（与列表新列布局一致，设计书决策 #11 朴素表格）
    FA_EXPORT_HEADERS = [
        '序号', '分公司', '内部编号', '品目编号', '品目名称', '规格',
        '序列号', '当前状态', '使用人', '部门', '入库日期', '供应商', '采购日期',
    ]

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """实例表导出：新列布局（品目联查 + 出生行派生）。"""
        import openpyxl
        from django.http import HttpResponse

        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '固定资产实例'
        ws.append(self.FA_EXPORT_HEADERS)

        serializer = FixedAssetSerializer(queryset, many=True)
        for idx, data in enumerate(serializer.data, start=1):
            ws.append([
                idx,
                data['branch_name'] or '',
                data['内部编号'],
                data['item_code'],
                data['item_name'],
                data['item_spec'],
                data['序列号'] or '待补录',
                data['当前状态'],
                data['使用人'],
                data['department_name'] or '',
                str(data['入库日期']) if data['入库日期'] else '',
                data['供应商'] or '',
                str(data['采购日期']) if data['采购日期'] else '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="fixed_assets.xlsx"'
        return response

    @action(detail=True, methods=['patch'], url_path='supplement')
    def supplement(self, request, pk=None):
        """序列号补录：仅 序列号/备注 两字段（manage_instances 权限，渐进录入）。"""
        instance = self.get_object()
        serializer = FixedAssetSupplementSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.序列号 = serializer.validated_data.get('序列号', '')
        instance.备注 = serializer.validated_data.get('备注', '')
        instance.save(update_fields=['序列号', '备注', 'updated_at'])
        return Response(FixedAssetSerializer(instance).data)

    @action(detail=True, methods=['get'], url_path='timeline')
    def timeline(self, request, pk=None):
        """实例生平：档案 + 出生行派生 + 关联全部明细行倒序（P2 验收：任一实例可查完整生平）。"""
        instance = self.get_object()

        birth = None
        if instance.birth_line is not None:
            birth_transfer = instance.birth_line.transfer
            birth = {
                'transfer_id': str(birth_transfer.pk),
                '单据编号': birth_transfer.单据编号,
                '日期': birth_transfer.调拨日期,
                '供应商': birth_transfer.供应商 or '',
                '单价': instance.birth_line.单价,
                '采购日期': birth_transfer.调拨日期,
            }

        rows = []
        links = instance.line_links.select_related(
            'line__transfer', 'line__item', 'line__department',
        ).order_by('-line__created_at')
        for link in links:
            line = link.line
            transfer = line.transfer
            rows.append({
                'transfer_id': str(transfer.pk),
                '单据编号': transfer.单据编号,
                'action_type': transfer.action_type,
                '日期': transfer.调拨日期,
                '行号': line.行号,
                '品目编号': line.item.asset_code,
                '数量': line.数量,
                '使用人': line.使用人,
                '部门': line.department.name if line.department else '',
                '本批规格': line.本批规格,
                '审批状态': transfer.审批状态,
            })

        return Response({
            'instance': FixedAssetSerializer(instance).data,
            'birth': birth,
            'timeline': rows,
        })

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='import')
    def import_excel(self, request):
        """实例 Excel 导入冻结（P2 第二刀）：绕过单据直写实例违反铁律，存量由迁移承载。"""
        return Response(
            {'detail': '实例导入已下线：存量实例由系统迁移承载，新增实例请走采购入库单'},
            status=status.HTTP_410_GONE,
        )
