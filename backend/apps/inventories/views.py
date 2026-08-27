from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from core.pagination import StandardPagination
from core.permissions import DataScopeMixin, validate_branches_in_scope
from apps.permissions.permissions import OperationPermission
from apps.audit.decorators import audit_log
from .models import InventoryTask, InventoryItem, InventoryInstanceItem, InventoryCheck
from .serializers import (
    InventoryTaskSerializer,
    InventoryItemSerializer,
    InventoryInstanceItemSerializer,
    InventoryCheckSerializer,
    CheckItemSerializer,
    CheckInstanceSerializer,
    RejectSerializer,
    RecountSerializer,
)
from .filters import InventoryTaskFilterSet
from .services import generate_variance_adjustments, task_target_column


class InventoryTaskViewSet(DataScopeMixin, viewsets.ModelViewSet):
    queryset = InventoryTask.objects.select_related(
        'branch', 'category', 'created_by', 'rejected_by',
    ).all()
    serializer_class = InventoryTaskSerializer
    filterset_class = InventoryTaskFilterSet
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    scope_branch_field = 'branch'
    # 审批 / 驳回要求 approve_inventory
    required_operations = {
        'approve': 'approve_inventory',
        'reject': 'approve_inventory',
    }

    def get_queryset(self):
        qs = super().get_queryset()
        return self.get_scoped_queryset(qs)

    def perform_create(self, serializer):
        # 越权校验：盘点目标分公司必须在操作者授权范围内（admin 豁免）
        validate_branches_in_scope(self.request.user, serializer.validated_data.get('branch'))
        serializer.save(created_by=self.request.user)

    def _transition(self, pk, target_status, before_save=None, **field_updates):
        """状态机转换：事务内锁定任务行 + 二次状态校验后应用转换，防并发竞态
        （如双 approve 导致库存双扣）。before_save(task) 在锁内、字段写入前调用，
        用于需持锁的副作用（库存调整 / 生成盘点项等）；其内 raise ValidationError
        会回滚事务并返回 400。返回 (task, error_response)，后者非 None 时直接返回。
        """
        from django.db import transaction
        with transaction.atomic():
            task = InventoryTask.objects.select_for_update().get(pk=pk)
            if not task.can_transition(target_status):
                return task, Response(
                    {'detail': f'当前状态「{task.get_status_display()}」无法执行此操作'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if before_save is not None:
                before_save(task)
            for field, value in field_updates.items():
                setattr(task, field, value)
            task.save()
            return task, None

    # ---- State transition actions ----

    @action(detail=True, methods=['post'])
    @audit_log(action='start', resource_type='InventoryTask', description_template='开始盘点')
    def start(self, request, pk=None):
        """开始盘点: pending -> in_progress"""
        task = self.get_object()
        if not task.can_transition('in_progress'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 开始盘点'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _check_and_prep(t):
            # 锁内复查同分公司是否已有进行中盘点，缩小并发窗口
            if t.branch and InventoryTask.objects.filter(
                branch=t.branch, status__in=['in_progress', 'pending_review'],
            ).exclude(pk=t.pk).exists():
                from rest_framework.exceptions import ValidationError
                raise ValidationError(
                    {'detail': f'分公司「{t.branch.name}」已有进行中的盘点任务，不可同时盘点'}
                )
            t.started_at = timezone.now()

        task, err = self._transition(
            pk, 'in_progress', before_save=_check_and_prep, status='in_progress',
        )
        if err:
            return err
        # 按任务类型生成清单：实例盘=部门在用实例快照；台账盘=范围内台账行
        if task.is_instance_inventory:
            self._generate_instance_items(task)
        else:
            self._generate_items(task)
        return Response(InventoryTaskSerializer(task).data)

    @action(detail=True, methods=['post'])
    def check(self, request, pk=None):
        """盘点单项"""
        task = self.get_object()
        if task.is_instance_inventory:
            return Response(
                {'detail': '实例盘任务请使用逐台核对（check-instance）'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if task.status != 'in_progress':
            return Response(
                {'detail': '只有盘点中的任务可以盘点'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CheckItemSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        from apps.assets.models import AssetStock
        stock = AssetStock.objects.select_related('item').filter(
            id=serializer.validated_data['stock_id'], branch=task.branch,
        ).first()
        if not stock:
            return Response(
                {'detail': '台账行不存在或不属于本盘点任务分公司'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Get or create the inventory item（应盘数量=任务库别对应列）
        column = task_target_column(task)
        item, _ = InventoryItem.objects.get_or_create(
            task=task, stock=stock,
            defaults={'expected_qty': getattr(stock, column)},
        )

        qty = serializer.validated_data['qty']
        remarks = serializer.validated_data.get('remarks', '')

        # Apply repeat rule
        if task.repeat_rule == 'last':
            item.actual_qty = qty
        else:  # accumulate
            item.actual_qty = (item.actual_qty or 0) + qty

        item.check_count += 1
        item.checked_by = request.user
        item.checked_at = timezone.now()
        item.remarks = remarks

        # Determine result
        if item.actual_qty == item.expected_qty:
            item.result = 'matched'
        elif item.actual_qty > item.expected_qty:
            item.result = 'surplus'
        else:
            item.result = 'missing'
        item.save()

        # Create check record
        check_record = InventoryCheck.objects.create(
            task=task, item=item, stock=stock, qty=qty,
            checked_by=request.user,
        )
        return Response(InventoryCheckSerializer(check_record).data)

    @action(detail=True, methods=['post'], url_path='check-instance')
    def check_instance(self, request, pk=None):
        """实例盘逐台核对：found=true→已找到 / false→未找到（重复核对以最后一次为准）"""
        task = self.get_object()
        if not task.is_instance_inventory:
            return Response(
                {'detail': '该任务不是部门实例盘点'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if task.status != 'in_progress':
            return Response(
                {'detail': '只有盘点中的任务可以核对'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CheckInstanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        entry = (
            task.instance_items
            .select_related('instance__item', 'instance__department')
            .filter(instance_id=serializer.validated_data['instance_id'])
            .first()
        )
        if not entry:
            return Response(
                {'detail': '实例不在本盘点任务清单内'},
                status=status.HTTP_404_NOT_FOUND,
            )

        found = serializer.validated_data['found']
        entry.result = 'matched' if found else 'missing'
        entry.check_count += 1
        entry.checked_by = request.user
        entry.checked_at = timezone.now()
        remarks = serializer.validated_data.get('remarks', '')
        if remarks:
            entry.remarks = remarks
        entry.save()
        return Response(InventoryInstanceItemSerializer(entry).data)

    @action(detail=True, methods=['post'])
    @audit_log(action='submit', resource_type='InventoryTask', description_template='提交盘点审核')
    def submit(self, request, pk=None):
        """提交审核: in_progress -> pending_review"""
        task = self.get_object()
        if not task.can_transition('pending_review'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 提交审核'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _prep(t):
            # Apply missed rule to unchecked items before submission
            self._apply_missed_rule(t)
            t.submitted_at = timezone.now()

        task, err = self._transition(
            pk, 'pending_review', before_save=_prep, status='pending_review',
        )
        if err:
            return err
        return Response(InventoryTaskSerializer(task).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, OperationPermission])
    @audit_log(action='approve', resource_type='InventoryTask', description_template='审批盘点任务')
    def approve(self, request, pk=None):
        """审核通过: pending_review -> completed（锁任务行 + 二次校验，防并发双扣库存）"""
        task = self.get_object()
        if not task.can_transition('completed'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 审核通过'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _adjust(t):
            # 差异自动生成调整单（P3，仅台账盘）：锁内事务逐差异项经唯一写入口修账，
            # 任一行致负数抛错整笔回滚（任务留 pending_review），差异不吞。
            # 实例盘不自动改账（盘亏报告待跟进，人工决定后续）
            if not t.is_instance_inventory:
                generate_variance_adjustments(t, request.user)
            t.completed_at = timezone.now()

        task, err = self._transition(
            pk, 'completed', before_save=_adjust, status='completed',
        )
        if err:
            return err
        return Response(InventoryTaskSerializer(task).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, OperationPermission])
    @audit_log(action='reject', resource_type='InventoryTask', description_template='驳回盘点任务')
    def reject(self, request, pk=None):
        """审核驳回: pending_review -> rejected"""
        task = self.get_object()
        if not task.can_transition('rejected'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 驳回'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = RejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        task, err = self._transition(
            pk, 'rejected',
            rejected_at=timezone.now(),
            rejected_by=request.user,
            reject_reason=serializer.validated_data['reason'],
            status='rejected',
        )
        if err:
            return err
        return Response(InventoryTaskSerializer(task).data)

    @action(detail=True, methods=['post'])
    def recount(self, request, pk=None):
        """重新盘点（驳回后）: rejected -> in_progress"""
        task = self.get_object()
        if not task.can_transition('in_progress'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 重新盘点'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = RecountSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reset_scope = serializer.validated_data.get('reset_scope', 'all')

        def _reset(t):
            if t.is_instance_inventory:
                qs = t.instance_items.all() if reset_scope == 'all' else (
                    t.instance_items.filter(result__in=['missing', 'unchecked'])
                )
                qs.update(
                    result='unchecked', check_count=0,
                    checked_by=None, checked_at=None,
                )
                return
            if reset_scope == 'abnormal_only':
                # Only reset abnormal items (surplus, missing, unchecked)
                t.items.filter(result__in=['surplus', 'missing', 'unchecked']).update(
                    actual_qty=None, result='unchecked',
                    check_count=0, checked_by=None, checked_at=None,
                )
            else:
                # Reset all items (original behavior)
                t.items.all().update(
                    actual_qty=None, result='unchecked',
                    check_count=0, checked_by=None, checked_at=None,
                )

        task, err = self._transition(
            pk, 'in_progress', before_save=_reset, status='in_progress',
        )
        if err:
            return err
        return Response(InventoryTaskSerializer(task).data)

    @action(detail=True, methods=['post'])
    @audit_log(action='cancel', resource_type='InventoryTask', description_template='作废盘点任务')
    def cancel(self, request, pk=None):
        """作废: pending/in_progress/rejected -> cancelled"""
        task = self.get_object()
        if not task.can_transition('cancelled'):
            return Response(
                {'detail': f'无法从 {task.get_status_display()} 作废'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        task, err = self._transition(pk, 'cancelled', status='cancelled')
        if err:
            return err
        return Response(InventoryTaskSerializer(task).data)

    # ---- Read-only endpoints ----

    def _build_progress_data(self, task, items=None):
        """Build progress stats dict with rates（台账盘/实例盘同口径，实例盘单位为台）."""
        if task.is_instance_inventory:
            if items is None:
                items = task.instance_items.all()
            total = items.count()
            matched = items.filter(result='matched').count()
            missing = items.filter(result='missing').count()
            unchecked = items.filter(result='unchecked').count()
            checked = matched + missing
            return {
                'totalItems': total,
                'checkedItems': checked,
                'matchedCount': matched,
                'surplusCount': 0,
                'missingCount': missing,
                'uncheckedCount': unchecked,
                'matchRate': round(matched / checked * 100, 1) if checked else 0,
                'surplusRate': 0,
                'missingRate': round(missing / checked * 100, 1) if checked else 0,
            }
        if items is None:
            items = task.items.all()
        total = items.count()
        checked = items.filter(result__in=['matched', 'surplus', 'missing']).count()
        matched = items.filter(result='matched').count()
        surplus = items.filter(result='surplus').count()
        missing = items.filter(result='missing').count()
        unchecked = items.filter(result='unchecked').count()
        return {
            'totalItems': total,
            'checkedItems': checked,
            'matchedCount': matched,
            'surplusCount': surplus,
            'missingCount': missing,
            'uncheckedCount': unchecked,
            'matchRate': round(matched / checked * 100, 1) if checked else 0,
            'surplusRate': round(surplus / checked * 100, 1) if checked else 0,
            'missingRate': round(missing / checked * 100, 1) if checked else 0,
        }

    @action(detail=True, methods=['get'])
    def progress(self, request, pk=None):
        """盘点进度"""
        task = self.get_object()
        return Response(self._build_progress_data(task))

    @action(detail=True, methods=['get'])
    def report(self, request, pk=None):
        """盘点报告（台账盘含差异调整单汇总；实例盘为实例清单与缺失明细）"""
        task = self.get_object()
        if task.is_instance_inventory:
            items = task.instance_items.select_related(
                'instance__item', 'instance__department', 'checked_by',
            ).all()
            return Response({
                'task': InventoryTaskSerializer(task).data,
                'progress': self._build_progress_data(task, items),
                'items': InventoryInstanceItemSerializer(items, many=True).data,
                'adjustments': {'total': 0, 'surplus': 0, 'missing': 0},
            })
        items = task.items.select_related('stock__item', 'stock__branch').all()
        adjustments = task.adjustments.all()
        data = {
            'task': InventoryTaskSerializer(task).data,
            'progress': self._build_progress_data(task, items),
            'items': InventoryItemSerializer(items, many=True).data,
            'adjustments': {
                'total': adjustments.count(),
                'surplus': adjustments.filter(变动量__gt=0).count(),
                'missing': adjustments.filter(变动量__lt=0).count(),
            },
        }
        return Response(data)

    @action(detail=True, methods=['get'])
    def checks(self, request, pk=None):
        """盘点记录（多人协作，可按盘点人 checkedBy 过滤）"""
        task = self.get_object()
        queryset = task.checks.select_related('stock__item', 'checked_by').all()

        checked_by = request.query_params.get('checkedBy')
        if checked_by:
            import uuid as uuid_lib
            try:
                checked_by = str(uuid_lib.UUID(checked_by))
            except (ValueError, AttributeError, TypeError):
                return Response(
                    {'detail': 'checkedBy 参数格式错误'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            queryset = queryset.filter(checked_by=checked_by)

        # Paginate
        paginator = StandardPagination()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = InventoryCheckSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)

        serializer = InventoryCheckSerializer(queryset, many=True)
        return Response(serializer.data)

    # ---- Excel import/export ----

    @action(detail=True, methods=['get'], url_path='export-report')
    def export_report(self, request, pk=None):
        """导出盘点报告 Excel：基本信息 + 统计 + 调整单清单 + 明细（双 sheet，任意状态可导）"""
        import io
        import openpyxl
        from django.http import HttpResponse

        task = self.get_object()

        def _dt(value):
            return value.strftime('%Y-%m-%d %H:%M') if value else ''

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '盘点报告'

        ws.append(['盘点报告'])
        ws.append([])
        ws.append(['基本信息'])
        base_info = [
            ('任务名称', task.name),
            ('分公司', task.branch.name if task.branch else '全部分公司'),
            ('资产类目', task.category.asset_category if task.category else '全部类目'),
        ]
        if task.is_instance_inventory:
            base_info.append(('盘点部门', task.department.name if task.department else ''))
        else:
            base_info.append(('库别', task.get_stock_bin_display()))
        base_info += [
            ('状态', task.get_status_display()),
            ('创建人', task.created_by.name if task.created_by else ''),
            ('创建时间', _dt(task.created_at)),
            ('开始时间', _dt(task.started_at)),
            ('提交时间', _dt(task.submitted_at)),
            ('完成时间', _dt(task.completed_at)),
            ('漏盘规则', task.get_missed_rule_display()),
        ]
        if not task.is_instance_inventory:
            base_info.append(('重复盘点规则', task.get_repeat_rule_display()))
        for label, value in base_info:
            ws.append([label, value])

        ws.append([])
        ws.append(['盘点统计'])
        if task.is_instance_inventory:
            items = task.instance_items.select_related(
                'instance__item', 'instance__department', 'checked_by',
            ).order_by('created_at')
            progress = self._build_progress_data(task, items)
            for label, value in [
                ('应到（台）', progress['totalItems']),
                ('实到（台）', progress['matchedCount']),
                ('缺失（台）', progress['missingCount']),
                ('未核对（台）', progress['uncheckedCount']),
                ('实到率', f"{progress['matchRate']}%"),
            ]:
                ws.append([label, value])
        else:
            items = task.items.select_related(
                'stock__item', 'stock__branch', 'checked_by',
            ).order_by('created_at')
            progress = self._build_progress_data(task, items)
            for label, value in [
                ('应盘品目数', progress['totalItems']),
                ('已盘品目数', progress['checkedItems']),
                ('正常', progress['matchedCount']),
                ('盘盈', progress['surplusCount']),
                ('盘亏', progress['missingCount']),
                ('未盘', progress['uncheckedCount']),
                ('正常率', f"{progress['matchRate']}%"),
                ('盘盈率', f"{progress['surplusRate']}%"),
                ('盘亏率', f"{progress['missingRate']}%"),
            ]:
                ws.append([label, value])

        ws.append([])
        ws.append(['差异调整单'])
        if task.is_instance_inventory:
            ws.append(['无（实例盘不改账）'])
        else:
            adjustments = task.adjustments.select_related('经办人').order_by('created_at')
            if adjustments:
                ws.append(['单据编号', '目标列', '变动量', '事由', '经办人', '创建时间'])
                for adj in adjustments:
                    ws.append([
                        adj.单据编号 or '',
                        adj.目标列,
                        adj.变动量,
                        adj.事由,
                        adj.经办人.name if adj.经办人 else '',
                        _dt(adj.created_at),
                    ])
            elif task.status == 'completed':
                ws.append(['无差异调整单'])
            else:
                ws.append(['无（任务未完成）'])

        ws_detail = wb.create_sheet('盘点明细')
        if task.is_instance_inventory:
            RESULT_LABELS = {'matched': '已找到', 'missing': '未找到（待跟进）', 'unchecked': '未核对'}
            ws_detail.append([
                '序号', '内部编号', '资产编号', '资产名称', '序列号',
                '使用人', '部门', '核对结果', '核对人', '核对时间', '备注',
            ])
            for idx, entry in enumerate(items, start=1):
                inst = entry.instance
                ws_detail.append([
                    idx,
                    inst.内部编号,
                    inst.item.asset_code,
                    inst.item.asset_name,
                    inst.序列号 or '（待补录）',
                    inst.使用人,
                    inst.department.name if inst.department else '',
                    RESULT_LABELS.get(entry.result, entry.result),
                    entry.checked_by.name if entry.checked_by else '',
                    _dt(entry.checked_at),
                    entry.remarks or '',
                ])
        else:
            ws_detail.append([
                '序号', '资产编号', '资产名称', '资产类目',
                '应盘', '实盘', '差异', '结果', '盘点人', '盘点时间', '备注',
            ])
            for idx, entry in enumerate(items, start=1):
                stock = entry.stock
                actual = entry.actual_qty
                diff = None if actual is None else actual - entry.expected_qty
                ws_detail.append([
                    idx,
                    stock.item.asset_code,
                    stock.item.asset_name,
                    stock.item.asset_category,
                    entry.expected_qty,
                    actual if actual is not None else '',
                    f'+{diff}' if diff is not None and diff > 0 else (str(diff) if diff is not None else ''),
                    entry.get_result_display(),
                    entry.checked_by.name if entry.checked_by else '',
                    _dt(entry.checked_at),
                    entry.remarks or '',
                ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'盘点报告_{task.name}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='import-template')
    def download_template(self, request, pk=None):
        """下载盘点模板 Excel"""
        import io
        import openpyxl
        from django.http import HttpResponse

        task = self.get_object()
        if task.is_instance_inventory:
            return Response(
                {'detail': '实例盘任务请逐台核对（点选/扫码），不提供 Excel 导入模板'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        items = task.items.select_related('stock__item').order_by('created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '盘点表'
        ws.append(['序号', '资产编号', '资产名称', '资产类目', '账面数量', '实盘数量', '备注'])

        for idx, item in enumerate(items, start=1):
            stock = item.stock
            ws.append([
                idx,
                stock.item.asset_code,
                stock.item.asset_name,
                stock.item.asset_category,
                item.expected_qty,
                '',  # 实盘数量 - 用户填写
                '',  # 备注 - 用户填写
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'盘点表_{task.name}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['post'], url_path='import-result',
            parser_classes=[MultiPartParser])
    def import_result(self, request, pk=None):
        """导入盘点结果 Excel"""
        task = self.get_object()
        if task.is_instance_inventory:
            return Response(
                {'detail': '实例盘任务请逐台核对（点选/扫码），不支持 Excel 导入'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if task.status != 'in_progress':
            return Response(
                {'detail': '只有盘点中的任务可以导入'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        from core.upload_validation import (
            validate_excel_upload, validate_row_count, UploadValidationError,
        )
        try:
            validate_excel_upload(file)
        except UploadValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {'detail': f'文件解析失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            validate_row_count(ws)
        except UploadValidationError as e:
            wb.close()
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        imported = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            if not row or not row[1]:
                continue
            asset_code = str(row[1]).strip()
            actual_qty_raw = row[5] if len(row) > 5 else None
            remarks = str(row[6]).strip() if len(row) > 6 and row[6] else ''

            if actual_qty_raw is None:
                continue

            try:
                actual_qty = int(float(str(actual_qty_raw)))
            except (ValueError, TypeError):
                errors.append(f'第 {i} 行: 实盘数量格式错误 "{actual_qty_raw}"')
                continue

            # Find the inventory item by asset code
            try:
                item = task.items.select_related('stock__item').get(stock__item__asset_code=asset_code)
            except InventoryItem.DoesNotExist:
                errors.append(f'第 {i} 行: 资产编号 "{asset_code}" 不在盘点范围内')
                continue

            # Apply repeat rule
            if task.repeat_rule == 'last':
                item.actual_qty = actual_qty
            else:  # accumulate
                item.actual_qty = (item.actual_qty or 0) + actual_qty

            item.check_count += 1
            item.checked_by = request.user
            item.checked_at = timezone.now()
            if remarks:
                item.remarks = remarks

            # Determine result
            if item.actual_qty == item.expected_qty:
                item.result = 'matched'
            elif item.actual_qty > item.expected_qty:
                item.result = 'surplus'
            else:
                item.result = 'missing'
            item.save()

            # Create check record
            InventoryCheck.objects.create(
                task=task, item=item, stock=item.stock, qty=actual_qty,
                checked_by=request.user,
            )
            imported += 1

        return Response({
            'imported': imported,
            'errors': errors,
        })

    # ---- Helpers ----

    def _generate_items(self, task):
        """台账盘：从任务范围内台账行生成盘点项（目标库别列>0，应盘=该列）。"""
        from apps.assets.models import AssetStock
        column = task_target_column(task)
        qs = AssetStock.objects.select_related('item').filter(
            **{f'{column}__gt': 0},
        )
        if task.branch:
            qs = qs.filter(branch=task.branch)
        if task.category:
            qs = qs.filter(item__asset_category=task.category.asset_category)
        for stock in qs:
            InventoryItem.objects.get_or_create(
                task=task, stock=stock,
                defaults={'expected_qty': getattr(stock, column)},
            )

    def _generate_instance_items(self, task):
        """实例盘：生成部门名下在用实例快照（一台一行；仅实例管理品目有实例）。"""
        from apps.assets.models import FixedAsset
        qs = FixedAsset.objects.select_related('item', 'department').filter(
            当前状态=FixedAsset.STATUS_IN_USE,
            branch=task.branch,
            department=task.department,
        )
        if task.category:
            qs = qs.filter(item__asset_category=task.category.asset_category)
        for instance in qs:
            InventoryInstanceItem.objects.get_or_create(task=task, instance=instance)

    def _apply_missed_rule(self, task):
        """Apply missed rule to unchecked items before submission."""
        if task.is_instance_inventory:
            # 实例盘：zero→未核对记缺失（进报告待跟进）；keep→保持未核对
            if task.missed_rule == 'zero':
                task.instance_items.filter(result='unchecked').update(result='missing')
            return
        unchecked_items = task.items.filter(result='unchecked')
        if task.missed_rule == 'zero':
            # Zero out: set actual_qty to 0 and result to 'missing'
            unchecked_items.update(actual_qty=0, result='missing')
        # 'keep' rule: leave unchecked items as-is (no change needed)

