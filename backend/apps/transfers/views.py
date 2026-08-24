import io
from datetime import date
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from core.pagination import StandardPagination
from core.permissions import DataScopeMixin, validate_branches_in_scope
from apps.permissions.permissions import OperationPermission
from apps.audit.decorators import audit_log
from .models import Transfer, TransferLine
from .serializers import (
    TransferSerializer, TransferActionSerializer, ApproveSerializer,
)
from .services import generate_document_number
from .filters import TransferFilterSet

# Active inventory statuses that lock a branch's transfers
_INVENTORY_LOCKED_STATUSES = ['in_progress', 'pending_review']


def _build_lines(transfer, items):
    """按提交顺序赋行号批量建明细行（行号唯一约束守护）。"""
    TransferLine.objects.bulk_create([
        TransferLine(transfer=transfer, 行号=idx, **item)
        for idx, item in enumerate(items, start=1)
    ])


def _notify_created(transfer):
    """明细行齐备后显式触发创建通知（post_save 时行尚未建，信号侧会静默跳过）。"""
    from apps.notifications.signals import notify_transfer_created
    notify_transfer_created(transfer)


class TransferViewSet(DataScopeMixin, viewsets.ModelViewSet):
    queryset = Transfer.objects.select_related(
        'from_branch', 'to_branch',
    ).prefetch_related('lines__item', 'lines__department').all()
    serializer_class = TransferSerializer
    filterset_class = TransferFilterSet
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    # 调拨按「调出 / 调入」双向分公司过滤
    scope_transfer_fields = ('from_branch', 'to_branch')
    # 审批要求 approve_transfer；入库确认要求 manage_assets；其余读写无声明即放行
    # 业务发起（purchase/assign/return/transfer/recovery）对所有登录用户开放
    # （员工申请领用 / 采购），数据范围由 _create_action 内 validate_branches_in_scope 控制；
    # 审批 / 入库 / 导入需授权。
    required_operations = {
        'import_excel': 'manage_assets',
        'approve': 'approve_transfer',
        'warehouse': 'manage_assets',
    }

    def get_queryset(self):
        qs = super().get_queryset()
        return self.get_scoped_queryset(qs)

    def _check_inventory_lock(self, branch_name=None, branch_id=None):
        """Raise ValidationError if branch has active inventory tasks."""
        from apps.inventories.models import InventoryTask
        from apps.organizations.models import Branch
        branch = None
        if branch_id:
            try:
                branch = Branch.objects.get(id=branch_id)
            except Branch.DoesNotExist:
                return
        elif branch_name:
            try:
                branch = Branch.objects.get(name=branch_name)
            except Branch.DoesNotExist:
                return
        if not branch:
            return
        if InventoryTask.objects.filter(
            branch=branch,
            status__in=_INVENTORY_LOCKED_STATUSES,
        ).exists():
            raise ValidationError({
                'detail': f'分公司「{branch.name}」正在进行盘点，暂时无法进行此操作',
                'code': 'INVENTORY_LOCKED',
            })

    def _resolve_branches(self, data):
        """外键优先、文字名称兜底解析分公司；解析失败即报错（导入路径依赖）。"""
        from apps.organizations.models import Branch
        from_branch = data.pop('from_branch', None)
        to_branch = data.pop('to_branch', None)
        if not from_branch and data.get('调出分公司'):
            from_branch = Branch.objects.filter(name=data['调出分公司']).first()
            if from_branch is None:
                raise ValidationError({'detail': f'调出分公司「{data["调出分公司"]}」不存在'})
        if not to_branch and data.get('调入分公司'):
            to_branch = Branch.objects.filter(name=data['调入分公司']).first()
            if to_branch is None:
                raise ValidationError({'detail': f'调入分公司「{data["调入分公司"]}」不存在'})
        # 回填分公司名称：表单只传外键 id 时保证筛选/展示一致
        if from_branch and not data.get('调出分公司'):
            data['调出分公司'] = from_branch.name
        if to_branch and not data.get('调入分公司'):
            data['调入分公司'] = to_branch.name
        return from_branch, to_branch

    def _create_action(self, request, action_type):
        """Shared helper for the 5 action routes：单头 + items 明细行。"""
        serializer = TransferActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        items = data.pop('items')
        data['action_type'] = action_type
        if not data.get('创建人'):
            data['创建人'] = request.user.name or request.user.phone

        # 草稿：保存为「草稿」状态，不进入审批流
        if request.data.get('draft'):
            data['审批状态'] = '草稿'

        from_branch, to_branch = self._resolve_branches(data)

        # 写操作越权校验：调出 / 调入分公司必须在操作者授权范围内（admin 豁免）
        validate_branches_in_scope(request.user, from_branch, to_branch)

        # Check inventory lock on both source and target branches
        self._check_inventory_lock(branch_id=from_branch.id if from_branch else None)
        self._check_inventory_lock(branch_id=to_branch.id if to_branch else None)

        # 行内直接回收（明细/固定资产列表发起）：manage_assets 持有者创建即「已通过」并即时联动
        immediate_recovery = bool(request.data.get('immediate')) and action_type == Transfer.ACTION_RECOVERY
        if immediate_recovery and not request.user.can('manage_assets'):
            return Response(
                {'detail': '直接回收需要资产管理权限'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if immediate_recovery:
            data['审批状态'] = '已通过'
            data['审批人'] = request.user.name or request.user.phone
            data['审批时间'] = timezone.now()

        from django.db import transaction
        with transaction.atomic():
            transfer = Transfer(
                单据编号=generate_document_number(action_type, data['调拨日期']),
                from_branch=from_branch,
                to_branch=to_branch,
                **data,
            )
            transfer.save()
            _build_lines(transfer, items)
            if immediate_recovery:
                self._apply_ledger(transfer)
        _notify_created(transfer)
        return Response(
            TransferSerializer(transfer).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'])
    @audit_log(action='purchase', resource_type='Transfer', description_template='采购入库')
    def purchase(self, request):
        """采购入库"""
        return self._create_action(request, Transfer.ACTION_PURCHASE)

    @action(detail=False, methods=['post'])
    @audit_log(action='assign', resource_type='Transfer', description_template='资产领用')
    def assign(self, request):
        """资产领用"""
        return self._create_action(request, Transfer.ACTION_ASSIGN)

    @action(detail=False, methods=['post'], url_path='return')
    @audit_log(action='return', resource_type='Transfer', description_template='资产归还')
    def return_asset(self, request):
        """资产归还 - mapped as 'return' on the frontend."""
        return self._create_action(request, Transfer.ACTION_RETURN)

    @action(detail=False, methods=['post'])
    @audit_log(action='transfer', resource_type='Transfer', description_template='资产调拨')
    def transfer(self, request):
        """资产调拨"""
        return self._create_action(request, Transfer.ACTION_TRANSFER)

    @action(detail=False, methods=['post'])
    @audit_log(action='recovery', resource_type='Transfer', description_template='资产回收')
    def recovery(self, request):
        """资产回收"""
        return self._create_action(request, Transfer.ACTION_RECOVERY)

    def _apply_ledger(self, transfer):
        """台账联动（唯一写入口 services/ledger.py 按明细行迭代）+ 回收按行内部编号删固定资产实例。

        FixedAsset 物理删除为 P1 过渡现状（P2 实例层接入后改为状态退役，设计书 5.3）。
        """
        from apps.assets.services import ledger

        ledger.apply_document(transfer)
        if transfer.action_type == Transfer.ACTION_RECOVERY:
            from apps.assets.models import FixedAsset
            inner_codes = [
                (line.固定资产内部编号 or '').strip()
                for line in transfer.lines.all()
            ]
            for inner_code in inner_codes:
                if inner_code:
                    FixedAsset.objects.filter(内部编号=inner_code).delete()

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, OperationPermission])
    @audit_log(action='approve', resource_type='Transfer', description_template='审批流转单')
    def approve(self, request, pk=None):
        """审批调拨单（事务内加锁，保证审批与资产同步原子、并发幂等）"""
        from django.db import transaction

        transfer = self.get_object()
        serializer = ApproveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        if transfer.审批状态 != '待审批':
            return Response(
                {'detail': '该记录已审批'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        approved = serializer.validated_data['approved']
        reason = serializer.validated_data.get('reason', '')

        with transaction.atomic():
            # 加锁重取，防止并发重复审批导致台账被多次联动
            locked = Transfer.objects.select_for_update().get(pk=transfer.pk)
            if locked.审批状态 != '待审批':
                return Response(
                    {'detail': '该记录已审批'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if approved:
                # 五单对称联动：充足性校验与数量变动均在唯一写入口内完成（按明细行迭代），
                # 不足时抛 ValidationError（LEDGER_INSUFFICIENT）→ 400 并整体回滚
                locked.审批状态 = (
                    '已入库' if locked.action_type == Transfer.ACTION_PURCHASE else '已通过'
                )
                self._apply_ledger(locked)
            else:
                locked.审批状态 = '已驳回'

            locked.审批人 = request.user.name or request.user.phone
            locked.审批时间 = timezone.now()
            if reason:
                locked.备注 = (locked.备注 + '\n' + reason).strip()
            locked.save(update_fields=[
                '审批状态', '审批人', '审批时间', '备注', 'updated_at',
            ])
        return Response(TransferSerializer(locked).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, OperationPermission])
    @audit_log(action='submit', resource_type='Transfer', description_template='提交采购草稿')
    def submit(self, request, pk=None):
        """提交草稿：将「草稿」转为「待审批」进入审批流。"""
        transfer = self.get_object()
        if transfer.审批状态 != '草稿':
            return Response({'detail': '仅草稿可提交'}, status=status.HTTP_400_BAD_REQUEST)
        transfer.审批状态 = '待审批'
        transfer.save(update_fields=['审批状态', 'updated_at'])
        _notify_created(transfer)
        return Response(TransferSerializer(transfer).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, OperationPermission])
    @audit_log(action='resubmit', resource_type='Transfer', description_template='重新提交流转')
    def resubmit(self, request, pk=None):
        """重新提交：将「已驳回」转为「待审批」重新进入审批流。"""
        transfer = self.get_object()
        if transfer.审批状态 != '已驳回':
            return Response({'detail': '仅已驳回的记录可重新提交'}, status=status.HTTP_400_BAD_REQUEST)
        transfer.审批状态 = '待审批'
        transfer.save(update_fields=['审批状态', 'updated_at'])
        _notify_created(transfer)
        return Response(TransferSerializer(transfer).data)

    def update(self, request, *args, **kwargs):
        """已驳回单据编辑：单头字段更新 + items 整体替换（原子、行号重排）。

        只应用请求里实际携带的字段——序列化器的字段默认值不得在 PATCH 时清空既有值。
        """
        from django.db import transaction

        kwargs.pop('partial', False)
        instance = self.get_object()
        if instance.审批状态 != '已驳回':
            return Response({'detail': '仅已驳回的记录可编辑'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = TransferActionSerializer(data=request.data, partial=True, for_update=True)
        serializer.is_valid(raise_exception=True)
        provided = set(request.data.keys())
        data = {k: v for k, v in serializer.validated_data.items() if k in provided}
        items = data.pop('items', None)

        from_branch = data.pop('from_branch', None)
        to_branch = data.pop('to_branch', None)
        if not from_branch and data.get('调出分公司'):
            from apps.organizations.models import Branch
            from_branch = Branch.objects.filter(name=data['调出分公司']).first()
            if from_branch is None:
                return Response(
                    {'detail': f'调出分公司「{data["调出分公司"]}」不存在'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        if not to_branch and data.get('调入分公司'):
            from apps.organizations.models import Branch
            to_branch = Branch.objects.filter(name=data['调入分公司']).first()
            if to_branch is None:
                return Response(
                    {'detail': f'调入分公司「{data["调入分公司"]}」不存在'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 越权校验：编辑后的调出 / 调入分公司必须在操作者授权范围内
        validate_branches_in_scope(
            request.user,
            from_branch or instance.from_branch,
            to_branch or instance.to_branch,
        )

        with transaction.atomic():
            for field, value in data.items():
                setattr(instance, field, value)
            if from_branch is not None:
                instance.from_branch = from_branch
            if to_branch is not None:
                instance.to_branch = to_branch
            instance.save()
            if items is not None:
                instance.lines.all().delete()
                _build_lines(instance, items)
        return Response(TransferSerializer(instance).data)

    def perform_update(self, serializer):
        raise AssertionError('update 已整体重写，不应走到默认 perform_update')

    def perform_destroy(self, instance):
        """已生效单据是对账流水的事实源，禁删；草稿/待审批/已驳回可删。"""
        if instance.审批状态 in ('已通过', '已入库'):
            raise ValidationError({'detail': '已生效单据不可删除（台账流水的 fact 来源）'})
        instance.delete()

    ACTION_TYPE_MAP = {
        '采购入库': Transfer.ACTION_PURCHASE,
        '领用': Transfer.ACTION_ASSIGN,
        '领用出库': Transfer.ACTION_ASSIGN,
        '归还': Transfer.ACTION_RETURN,
        '调拨': Transfer.ACTION_TRANSFER,
        '回收': Transfer.ACTION_RECOVERY,
    }

    TYPE_TEMPLATES = {
        'purchase': {
            'headers': ['采购日期', '分公司', '资产编号', '物品名称', '规格型号', '图片',
                        '供应商', '采购数量', '单价', '总金额', '需求部门', '采购经办人', '备注'],
            'sheet': '采购入库',
            'filename': 'purchase_template.xlsx',
        },
        'assign': {
            'headers': ['分公司', '日期', '资产编号', '领用物品', '领用数量', '用途', '领用部门', '备注'],
            'sheet': '领用出库',
            'filename': 'assign_template.xlsx',
        },
        'transfer': {
            'headers': ['调拨日期', '调出分公司', '调出部门', '调入分公司', '调入部门',
                        '资产编号', '资产名称', '规格型号', '调拨数量', '调拨原因',
                        '调出负责人', '调入负责人', '备注'],
            'sheet': '调拨',
            'filename': 'transfer_template.xlsx',
        },
        'recovery': {
            'headers': ['分公司', '资产编号', '资产类目', '物品分类', '资产名称', '回收分类',
                        '入库日期', '数量', '单位', '规格', '出库日期', '所属部门',
                        '存放位置', '经办人', '备注'],
            'sheet': '回收',
            'filename': 'recovery_template.xlsx',
        },
    }

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """下载空白导入模板，按 type 参数区分。"""
        import openpyxl
        from django.http import HttpResponse

        template_type = request.query_params.get('type', 'transfer')
        tpl = self.TYPE_TEMPLATES.get(template_type, self.TYPE_TEMPLATES['transfer'])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tpl['sheet']
        ws.append(tpl['headers'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{tpl["filename"]}"'
        return response

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """导出流转记录 Excel：按明细行展开输出（单头信息随行重复），模板列不变。"""
        import openpyxl
        from django.http import HttpResponse
        from django.db.models import Count
        from apps.assets.models import AssetStock

        queryset = self.filter_queryset(self.get_queryset())
        template_type = request.query_params.get('type', 'transfer')

        wb = openpyxl.Workbook()
        ws = wb.active

        def _spec(line):
            return line.本批规格 or line.item.specification

        if template_type == 'purchase':
            ws.title = '采购入库'
            headers = ['采购日期', '分公司', '资产编号', '物品名称', '规格型号', '图片',
                       '供应商', '采购数量', '单价', '总金额', '需求部门', '采购经办人', '备注']
            ws.append(headers)
            for t in queryset:
                for line in t.lines.all():
                    ws.append([
                        str(t.调拨日期) if t.调拨日期 else '',
                        t.调出分公司, line.item.asset_code, line.item.asset_name, _spec(line), '',
                        t.供应商, line.数量, line.单价 or '', line.金额 or '',
                        t.需求部门, t.采购经办人, t.备注,
                    ])

        elif template_type == 'assign':
            ws.title = '领用出库'
            headers = ['分公司', '日期', '资产编号', '领用物品', '领用数量', '用途', '领用部门',
                       '部门累计领用', '当前库存', '是否核对', '备注']
            ws.append(headers)

            dept_counts = dict(
                Transfer.objects.filter(
                    action_type=Transfer.ACTION_ASSIGN,
                    审批状态__in=['已通过', '已入库'],
                ).values('调出分公司', '调出部门').annotate(cnt=Count('id')).values_list('调出分公司', '调出部门', 'cnt')
            )
            dept_count_map = {(k[0], k[1]): k[2] for k in dept_counts}
            stocks = AssetStock.objects.select_related('branch', 'item')
            stock_map = {(s.branch.name, s.item.asset_code): s.在库数量 for s in stocks}

            for t in queryset:
                dept_total = dept_count_map.get((t.调出分公司, t.调出部门), 0)
                for line in t.lines.all():
                    current_stock = stock_map.get((t.调出分公司, line.item.asset_code), 0)
                    ws.append([
                        t.调出分公司,
                        str(t.调拨日期) if t.调拨日期 else '',
                        line.item.asset_code, line.item.asset_name, line.数量,
                        t.用途, t.调出部门,
                        dept_total, current_stock, '待核对', t.备注,
                    ])

        elif template_type == 'recovery':
            ws.title = '回收'
            headers = ['序号', '分公司', '资产编号', '资产类目', '物品分类', '资产名称',
                       '回收分类', '入库日期', '数量', '单位', '规格', '出库日期',
                       '所属部门', '当前处理状态', '存放位置', '经办人', '备注']
            ws.append(headers)
            idx = 0
            for t in queryset:
                for line in t.lines.all():
                    idx += 1
                    ws.append([
                        idx, t.调出分公司, line.item.asset_code,
                        line.item.asset_category, line.item.item_category,
                        line.item.asset_name, t.回收分类,
                        str(t.调拨日期) if t.调拨日期 else '',
                        line.数量, line.item.unit, _spec(line),
                        str(t.出库日期) if t.出库日期 else '',
                        t.调出部门, t.审批状态, line.存放位置, t.采购经办人, t.备注,
                    ])

        else:
            ws.title = '调拨'
            headers = ['调拨日期', '调出分公司', '调出部门', '调入分公司', '调入部门',
                       '资产编号', '资产名称', '规格型号', '调拨数量', '调拨原因',
                       '调出负责人', '调入负责人', '备注']
            ws.append(headers)
            for t in queryset:
                for line in t.lines.all():
                    ws.append([
                        str(t.调拨日期) if t.调拨日期 else '',
                        t.调出分公司, t.调出部门, t.调入分公司, t.调入部门,
                        line.item.asset_code, line.item.asset_name, _spec(line), line.数量,
                        t.调拨原因, t.调出负责人, t.调入负责人, t.备注,
                    ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{template_type}'
            f'_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='import')
    def import_excel(self, request):
        """批量导入：一行 = 一张单头 + 一条明细行（模板列与校验口径不变）。"""
        import openpyxl
        from datetime import datetime as dt
        from django.db import transaction
        from apps.categories.models import Category
        from apps.categories.views import suggest_similar_codes
        from apps.organizations.models import Branch
        from apps.organizations.utils import get_branch_name_set, branch_validation_error

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        from core.upload_validation import (
            validate_excel_upload, validate_row_count, UploadValidationError,
        )
        try:
            validate_excel_upload(file)
        except UploadValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        template_type = request.query_params.get('type', 'transfer')

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {'detail': f'文件解析失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            validate_row_count(ws)
        except UploadValidationError as e:
            wb.close()
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        imported = 0
        errors = []
        creator = request.user.name or request.user.phone
        valid_branches = get_branch_name_set()
        branch_cache = {b.name: b for b in Branch.objects.all()}
        item_cache = {c.asset_code: c for c in Category.objects.all()}

        def _parse_date(val):
            if hasattr(val, 'strftime'):
                return val
            return dt.strptime(str(val).strip(), '%Y-%m-%d').date()

        def _cell(row, idx):
            return str(row[idx]).strip() if len(row) > idx and row[idx] is not None else ''

        for i, row in enumerate(rows, start=2):
            try:
                # 分公司存在性校验（按类型取相关列）
                if template_type == 'purchase':
                    _branch_err = branch_validation_error(_cell(row, 1), '调出分公司', valid_branches)
                elif template_type == 'assign':
                    _branch_err = branch_validation_error(_cell(row, 0), '调出分公司', valid_branches)
                elif template_type == 'recovery':
                    _branch_err = branch_validation_error(_cell(row, 0), '调出分公司', valid_branches)
                else:  # transfer
                    _branch_err = branch_validation_error(_cell(row, 1), '调出分公司', valid_branches) \
                        or branch_validation_error(_cell(row, 3), '调入分公司', valid_branches)
                if _branch_err:
                    errors.append(f'第 {i} 行: {_branch_err}')
                    continue

                # 编号户籍校验：导入行必须携带字典内资产编号
                if template_type == 'purchase':
                    _code_idx = 2
                elif template_type == 'assign':
                    _code_idx = 2
                elif template_type == 'recovery':
                    _code_idx = 1
                else:
                    _code_idx = 5
                _code = _cell(row, _code_idx)
                if not _code:
                    errors.append(f'第 {i} 行: 资产编号为空')
                    continue
                item = item_cache.get(_code)
                if item is None:
                    _similar = suggest_similar_codes(_code)
                    _hint = f'，是否想找：{"、".join(_similar)}' if _similar else ''
                    errors.append(f'第 {i} 行: 资产编号 {_code} 未在品目字典登记{_hint}')
                    continue

                def _qty(idx, default=1):
                    raw = row[idx] if len(row) > idx else None
                    return int(raw) if raw else default

                with transaction.atomic():
                    if template_type == 'purchase':
                        branch_name = _cell(row, 1)
                        header = {
                            '调拨日期': _parse_date(row[0]),
                            '调出分公司': branch_name,
                            '供应商': _cell(row, 6),
                            '需求部门': _cell(row, 10),
                            '采购经办人': _cell(row, 11),
                            '备注': _cell(row, 12),
                        }
                        line_kwargs = {
                            'item': item,
                            '数量': _qty(7),
                            '本批规格': _cell(row, 4),
                            '单价': row[8] if len(row) > 8 and row[8] is not None else None,
                            '金额': row[9] if len(row) > 9 and row[9] is not None else None,
                        }
                        action = Transfer.ACTION_PURCHASE

                    elif template_type == 'assign':
                        branch_name = _cell(row, 0)
                        header = {
                            '调拨日期': _parse_date(row[1]),
                            '调出分公司': branch_name,
                            '用途': _cell(row, 5),
                            '调出部门': _cell(row, 6),
                            '备注': _cell(row, 7),
                        }
                        line_kwargs = {'item': item, '数量': _qty(4)}
                        action = Transfer.ACTION_ASSIGN

                    elif template_type == 'recovery':
                        branch_name = _cell(row, 0)
                        header = {
                            '调拨日期': _parse_date(row[6]) if row[6] else date.today(),
                            '调出分公司': branch_name,
                            '回收分类': _cell(row, 5),
                            '出库日期': _parse_date(row[10]) if row[10] else None,
                            '调出部门': _cell(row, 11),
                            '采购经办人': _cell(row, 14),
                            '备注': _cell(row, 15),
                        }
                        line_kwargs = {
                            'item': item,
                            '数量': _qty(7),
                            '本批规格': _cell(row, 9),
                            '存放位置': _cell(row, 13),
                        }
                        action = Transfer.ACTION_RECOVERY

                    else:  # transfer
                        header = {
                            '调拨日期': _parse_date(row[0]),
                            '调出分公司': _cell(row, 1),
                            '调出部门': _cell(row, 2),
                            '调入分公司': _cell(row, 3),
                            '调入部门': _cell(row, 4),
                            '调拨原因': _cell(row, 9),
                            '调出负责人': _cell(row, 10),
                            '调入负责人': _cell(row, 11),
                            '备注': _cell(row, 12),
                        }
                        line_kwargs = {'item': item, '数量': _qty(8), '本批规格': _cell(row, 7)}
                        action = Transfer.ACTION_TRANSFER

                    transfer = Transfer(
                        action_type=action,
                        审批状态='待审批',
                        创建人=creator,
                        单据编号=generate_document_number(action, header['调拨日期']),
                        from_branch=branch_cache.get(header.get('调出分公司', '')),
                        to_branch=branch_cache.get(header.get('调入分公司', '')) if header.get('调入分公司') else None,
                        **header,
                    )
                    transfer.save()
                    TransferLine.objects.create(transfer=transfer, 行号=1, **line_kwargs)
                _notify_created(transfer)
                imported += 1
            except Exception as e:
                errors.append(f'第 {i} 行: {str(e)}')

        wb.close()
        return Response({'imported': imported, 'errors': errors})
