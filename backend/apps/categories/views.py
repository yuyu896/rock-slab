import difflib
import io
from django.db import IntegrityError
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from apps.permissions.permissions import OperationPermission
from core.pagination import StandardPagination
from .models import Category
from .serializers import CategorySerializer
from .filters import CategoryFilterSet


def suggest_similar_codes(code, limit=3):
    """按编号在字典中找相近编号，供「未登记编号」错误提示。"""
    codes = list(Category.objects.values_list('asset_code', flat=True))
    return difflib.get_close_matches(code, codes, n=limit, cutoff=0.5)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filterset_class = CategoryFilterSet
    permission_classes = [IsAuthenticated, OperationPermission]
    pagination_class = StandardPagination
    # 创建/更新/删除/导入需 manage_dictionary（品目字典）；读取无声明即放行
    required_operations = {
        'create': 'manage_dictionary',
        'update': 'manage_dictionary',
        'partial_update': 'manage_dictionary',
        'destroy': 'manage_dictionary',
        'import_excel': 'manage_dictionary',
    }

    @action(detail=False, methods=['get'], url_path='lookup')
    def lookup(self, request):
        """按资产编号精确查询单条品目，供表单失焦反查名称/类目/管理方式等。"""
        code = (request.query_params.get('asset_code') or '').strip()
        if not code:
            return Response(
                {'asset_code': ['请提供 asset_code 参数']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        category = Category.objects.filter(asset_code=code).first()
        if category is None:
            similar = suggest_similar_codes(code)
            detail = '该资产编号未在品目字典登记'
            if similar:
                detail += f'，是否想找：{"、".join(similar)}'
            return Response({'detail': detail}, status=status.HTTP_404_NOT_FOUND)
        return Response({
            'id': str(category.id),
            '资产名称': category.asset_name,
            '资产类目': category.asset_category,
            '物品分类': category.item_category,
            '规格': category.specification,
            '管理方式': category.management_type,
            '是否租用': category.is_rental,
            '默认供应商': category.default_supplier,
            '计量单位': category.unit,
            '警戒线': category.warning_line,
        })

    def destroy(self, request, *args, **kwargs):
        """删除保护：被台账/资产/固定资产/流转单明细行引用的品目禁止删除。"""
        instance = self.get_object()
        from apps.assets.models import AssetStock, FixedAsset
        from apps.transfers.models import TransferLine

        references = []
        if AssetStock.objects.filter(item=instance).exists():
            references.append('台账')
        if FixedAsset.objects.filter(item=instance).exists():
            references.append('固定资产实例')
        if TransferLine.objects.filter(item=instance).exists():
            references.append('流转单')
        if references:
            return Response(
                {'detail': f'该品目已被{"、".join(references)}引用，不能删除'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        """创建分类，处理重复编号错误"""
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except IntegrityError:
            return Response(
                {'asset_code': ['资产编号已存在，请使用其他编号']},
                status=status.HTTP_400_BAD_REQUEST
            )

    def update(self, request, *args, **kwargs):
        """更新分类，处理重复编号错误"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        try:
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except IntegrityError:
            return Response(
                {'asset_code': ['资产编号已存在，请使用其他编号']},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """导出分类数据为 Excel"""
        import openpyxl

        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '分类数据'

        headers = ['资产类目', '物品分类', '资产名称', '资产编号', '规格', '管理方式', '是否租用', '默认供应商', '计量单位', '警戒线', '备注']
        ws.append(headers)

        column_widths = [15, 15, 20, 15, 15, 10, 10, 18, 12, 10, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        for obj in queryset:
            ws.append([
                obj.asset_category,
                obj.item_category,
                obj.asset_name,
                obj.asset_code,
                obj.specification,
                obj.get_management_type_display(),
                '是' if obj.is_rental else '否',
                obj.default_supplier,
                obj.unit,
                obj.warning_line or '',
                obj.remarks or '',
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="categories_export.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """下载分类导入 Excel 模板"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '分类导入模板'

        headers = ['资产类目', '物品分类', '资产名称', '资产编号', '计量单位', '警戒线', '备注']
        ws.append(headers)

        # 设置列宽
        column_widths = [15, 15, 20, 15, 12, 10, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="category_import_template.xlsx"'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='import')
    def import_excel(self, request):
        """Excel 批量导入分类：sheet 名定管理方式。

        表名精确匹配 数量管理/实例管理/消耗品 之一的，整表按对应管理方式导入；
        全簿无命中表名时回退 active sheet 按数量管理导入（兼容旧单表文件）。
        更新已存在品目且管理方式变化时受存量守卫约束（见 management_stock_status）。
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        from core.upload_validation import (
            validate_excel_upload, DEFAULT_MAX_ROWS, UploadValidationError,
        )
        try:
            validate_excel_upload(file)
        except UploadValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
        except Exception as e:
            return Response(
                {'detail': f'文件解析失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sheet_mgmt_map = dict(Category.MANAGEMENT_CHOICES_LABELS)
        matched = [(ws, sheet_mgmt_map[ws.title]) for ws in wb.worksheets if ws.title in sheet_mgmt_map]
        if matched:
            targets = matched
            skipped_sheets = [ws.title for ws in wb.worksheets if ws.title not in sheet_mgmt_map]
            fallback = False
        else:
            targets = [(wb.active, 'quantity')]
            skipped_sheets = []
            fallback = True

        # 行数上限按命中表合计（防拆表绕限）；回退路径单表合计 = 原口径
        total_rows = sum(
            max(0, (getattr(ws, 'max_row', None) or 0) - 1) for ws, _ in targets
        )
        if total_rows > DEFAULT_MAX_ROWS:
            wb.close()
            return Response(
                {'detail': f'数据行数过多（{total_rows} 行），最大支持 {DEFAULT_MAX_ROWS} 行'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .serializers import management_stock_status
        total_imported = 0
        all_errors = []
        sheet_stats = []

        for ws, management_type in targets:
            imported = 0
            errors = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    asset_category = str(row[0] or '').strip()
                    item_category = str(row[1] or '').strip()
                    asset_name = str(row[2] or '').strip()
                    asset_code = str(row[3] or '').strip()
                    unit = str(row[4] or '').strip()

                    if not all([asset_category, item_category, asset_name, asset_code, unit]):
                        errors.append(f'第 {i} 行: 必填字段不能为空')
                        continue

                    existing = Category.objects.filter(asset_code=asset_code).first()
                    if existing and existing.management_type != management_type:
                        locked, reason = management_stock_status(existing)
                        if locked:
                            label = Category.MANAGEMENT_LABELS.get(management_type, management_type)
                            errors.append(
                                f'第 {i} 行: {asset_code} 已有存量（{reason}），'
                                f'管理方式不可改为「{label}」，已跳过'
                            )
                            continue

                    Category.objects.update_or_create(
                        asset_code=asset_code,
                        defaults={
                            'asset_category': asset_category,
                            'item_category': item_category,
                            'asset_name': asset_name,
                            'unit': unit,
                            'management_type': management_type,
                            'warning_line': int(row[5]) if row[5] else None,
                            'remarks': str(row[6] or ''),
                        },
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f'第 {i} 行: {str(e)}')

            sheet_stats.append({
                'name': ws.title,
                'management_type': management_type,
                'imported': imported,
                'errors': errors,
            })
            total_imported += imported
            all_errors.extend(errors)

        wb.close()
        return Response({
            'imported': total_imported,
            'errors': all_errors,
            'sheets': sheet_stats,
            'skipped_sheets': skipped_sheets,
            'fallback': fallback,
        })
