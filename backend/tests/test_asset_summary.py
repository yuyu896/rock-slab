"""台账（AssetStock V2）契约测试：行粒度、四列语义、只读接口、增量导入。

对应 ledger-single-source / asset-summary(MODIFIED) 能力。
"""
import io

import openpyxl
import pytest
from conftest import _client_for
from rest_framework import status

SUMMARY_URL = '/api/assets/summary'


def _get_item(code):
    from apps.categories.models import Category
    item, _ = Category.objects.get_or_create(
        asset_code=code,
        defaults={
            'asset_category': '测试类目', 'item_category': '测试分类',
            'asset_name': f'品目 {code}', 'unit': '个',
        },
    )
    return item


def _make_stock(branch, code, stock=10, in_use=0, recycle=0, warning=None):
    from apps.assets.models import AssetStock
    row, _ = AssetStock.objects.get_or_create(
        branch=branch, item=_get_item(code),
        defaults={'在库数量': stock, '在用数量': in_use, '回收库数量': recycle, '警戒线': warning},
    )
    return row


def _xlsx_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'test.xlsx'
    return buf


# ---------------------------------------------------------------------------
# 模型不变量
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAssetStockModel:
    def test_unique_branch_item(self, branch):
        from apps.assets.models import AssetStock
        from django.db import IntegrityError
        _make_stock(branch, 'SUM-U-001')
        with pytest.raises(IntegrityError):
            AssetStock.objects.create(branch=branch, item=_get_item('SUM-U-001'))

    def test_total_is_three_columns_sum(self, branch):
        row = _make_stock(branch, 'SUM-T-001', stock=10, in_use=3, recycle=2)
        assert row.总量 == 15
        assert row.在库数量 == 10

    def test_sufficiency_uses_row_warning_first(self, branch):
        row = _make_stock(branch, 'SUM-W-001', stock=8, warning=10)
        item = row.item
        item.warning_line = 5
        item.save()
        assert row.警戒线 == 10
        assert row.是否充足 is False  # 行级 10 优先于字典默认 5

    def test_sufficiency_falls_back_to_dictionary_default(self, branch):
        row = _make_stock(branch, 'SUM-W-002', stock=3, warning=None)
        item = row.item
        item.warning_line = 5
        item.save()
        row.警戒线 = None
        row.save()
        assert row.生效警戒线 == 5
        assert row.是否充足 is False

    def test_no_warning_line_means_sufficient(self, branch):
        row = _make_stock(branch, 'SUM-W-003', stock=0, warning=None)
        item = row.item
        item.warning_line = None
        item.save()
        assert row.是否充足 is True


# ---------------------------------------------------------------------------
# 列表接口
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAssetStockList:
    def test_list_paginated_shape_with_dictionary_join(self, authenticated_client, branch):
        _make_stock(branch, 'SUM-L-001', stock=7, in_use=2)
        resp = authenticated_client.get(SUMMARY_URL)
        assert resp.status_code == 200
        results = resp.data['results']
        assert len(results) == 1
        row = results[0]
        assert row['资产编号'] == 'SUM-L-001'
        assert row['资产名称'] == '品目 SUM-L-001'
        assert row['在库数量'] == 7
        assert row['在用数量'] == 2
        assert row['回收库数量'] == 0
        assert row['总量'] == 9
        assert '管理方式' in row

    def test_list_filters(self, authenticated_client, branch, second_branch):
        _make_stock(branch, 'SUM-F-001')
        _make_stock(second_branch, 'SUM-F-002')
        resp = authenticated_client.get(SUMMARY_URL, {'branch': branch.name})
        assert resp.status_code == 200
        codes = {r['资产编号'] for r in resp.data['results']}
        assert codes == {'SUM-F-001'}

    def test_list_keyword_matches_item_code_or_name(self, authenticated_client, branch):
        _make_stock(branch, 'SUM-K-001')
        resp = authenticated_client.get(SUMMARY_URL, {'keyword': 'SUM-K'})
        assert {r['资产编号'] for r in resp.data['results']} == {'SUM-K-001'}
        resp = authenticated_client.get(SUMMARY_URL, {'keyword': '不存在的关键词'})
        assert resp.data['results'] == []

    def test_list_sufficient_filter(self, authenticated_client, branch):
        # 行级警戒线不足 / 品目默认不足 / 充足 三行
        from apps.categories.models import Category
        _make_stock(branch, 'SUM-W-001', stock=2, warning=5)
        low_item = _get_item('SUM-W-002')
        Category.objects.filter(pk=low_item.pk).update(warning_line=10)
        _make_stock(branch, 'SUM-W-002', stock=3)
        ok_item = _get_item('SUM-W-003')
        Category.objects.filter(pk=ok_item.pk).update(warning_line=1)
        _make_stock(branch, 'SUM-W-003', stock=5)

        resp = authenticated_client.get(SUMMARY_URL, {'sufficient': '0'})
        assert {r['资产编号'] for r in resp.data['results']} == {'SUM-W-001', 'SUM-W-002'}

        resp = authenticated_client.get(SUMMARY_URL, {'sufficient': '1'})
        assert {r['资产编号'] for r in resp.data['results']} == {'SUM-W-003'}

    def test_list_scoped_to_authorized_branches(self, supervisor_user, branch, second_branch):
        _make_stock(branch, 'SUM-S-001')
        _make_stock(second_branch, 'SUM-S-002')
        client = _client_for(supervisor_user)
        resp = client.get(SUMMARY_URL)
        codes = {r['资产编号'] for r in resp.data['results']}
        assert 'SUM-S-001' in codes
        assert 'SUM-S-002' not in codes

    def test_list_unauthenticated(self, api_client):
        resp = api_client.get(SUMMARY_URL)
        assert resp.status_code == 401


# ---------------------------------------------------------------------------
# 行级写接口下线（铁律 2）
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAssetStockReadOnly:
    def test_create_returns_405(self, authenticated_client, branch, category):
        resp = authenticated_client.post(SUMMARY_URL, {
            'branch': str(branch.id), 'item': str(category.id), '在库数量': 5,
        }, format='json')
        assert resp.status_code == status.HTTP_405_METHOD_NOT_ALLOWED

    def test_update_and_delete_return_405(self, authenticated_client, branch):
        row = _make_stock(branch, 'SUM-R-001')
        resp = authenticated_client.patch(f'{SUMMARY_URL}/{row.id}', {'在库数量': 99}, format='json')
        assert resp.status_code == 405
        resp = authenticated_client.delete(f'{SUMMARY_URL}/{row.id}')
        assert resp.status_code == 405

    def test_batch_delete_returns_405(self, authenticated_client):
        resp = authenticated_client.post(f'{SUMMARY_URL}/batch-delete', {'ids': []}, format='json')
        assert resp.status_code == 405


# ---------------------------------------------------------------------------
# 增量导入（预览 → 确认生成调整单）
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAssetStockIncrementalImport:
    def _upload(self, client, rows, confirm=False):
        buf = _xlsx_bytes(['分公司', '资产编号', '在库数量'], rows)
        data = {'file': buf}
        if confirm:
            data['confirm'] = '1'
        return client.post(f'{SUMMARY_URL}/import', data, format='multipart')

    def test_preview_returns_diffs(self, supervisor_user, branch):
        _make_stock(branch, 'SUM-I-001', stock=10)
        _get_item('SUM-I-002')
        client = _client_for(supervisor_user)
        resp = self._upload(client, [[branch.name, 'SUM-I-001', 12], [branch.name, 'SUM-I-002', 5]])
        assert resp.status_code == 200
        diffs = resp.data['diffs']
        assert len(diffs) == 2
        d1 = next(d for d in diffs if d['资产编号'] == 'SUM-I-001')
        assert d1['现值'] == 10 and d1['导入值'] == 12 and d1['变动量'] == 2
        d2 = next(d for d in diffs if d['资产编号'] == 'SUM-I-002')
        assert d2['现值'] == 0 and d2['导入值'] == 5

    def test_confirm_generates_adjustments(self, supervisor_user, branch):
        from apps.assets.models import AssetStock, LedgerAdjustment
        _make_stock(branch, 'SUM-I-003', stock=10)
        client = _client_for(supervisor_user)
        resp = self._upload(client, [[branch.name, 'SUM-I-003', 12]], confirm=True)
        assert resp.status_code == 200
        assert resp.data['applied'] == 1
        row = AssetStock.objects.get(branch=branch, item__asset_code='SUM-I-003')
        assert row.在库数量 == 12
        adj = LedgerAdjustment.objects.get(item__asset_code='SUM-I-003')
        assert adj.变动量 == 2
        assert '导入调整' in adj.事由

    def test_unchanged_rows_skipped(self, supervisor_user, branch):
        from apps.assets.models import AssetStock
        _make_stock(branch, 'SUM-I-004', stock=10)
        client = _client_for(supervisor_user)
        resp = self._upload(client, [[branch.name, 'SUM-I-004', 10]], confirm=True)
        assert resp.data['applied'] == 0
        row = AssetStock.objects.get(branch=branch, item__asset_code='SUM-I-004')
        assert row.在库数量 == 10

    def test_unregistered_code_rejected_with_hint(self, supervisor_user, branch):
        _get_item('SUM-I-005')
        client = _client_for(supervisor_user)
        resp = self._upload(client, [[branch.name, 'SUM-I-005X', 3]])
        assert resp.status_code == 200
        assert any('未在品目字典登记' in e for e in resp.data['errors'])
        assert resp.data['diffs'] == []

    def test_import_requires_permission(self, staff_user, branch):
        client = _client_for(staff_user)
        resp = self._upload(client, [[branch.name, 'SUM-I-006', 1]])
        assert resp.status_code == 403

    def test_template_headers(self, supervisor_user):
        client = _client_for(supervisor_user)
        resp = client.get(f'{SUMMARY_URL}/template')
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == ['分公司', '资产编号', '在库数量']

    def test_export_columns(self, supervisor_client, branch):
        _make_stock(branch, 'SUM-E-001', stock=5, in_use=1)
        resp = supervisor_client.get(f'{SUMMARY_URL}/export')
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == [
            '序号', '分公司', '资产编号', '资产名称', '规格', '资产类目', '物品分类',
            '管理方式', '在库数量', '在用数量', '回收库数量', '总量', '警戒线', '是否充足',
        ]
        row2 = [c.value for c in ws[2]]
        assert row2[2] == 'SUM-E-001'
        assert row2[8] == 5 and row2[11] == 6
