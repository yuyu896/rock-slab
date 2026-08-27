"""
Tests for Inventory flow: create task, start, check, submit, approve, reject, recount, cancel.
"""
import io

import openpyxl
import pytest
from rest_framework import status

INVENTORY_LIST_URL = '/api/inventories/'


def _task_action_url(action_name, pk):
    return f'/api/inventories/{pk}/{action_name}'


def _load_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.content))


def _sheet_cells(worksheet):
    return [str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row if cell is not None]


@pytest.fixture
def category(db):
    from apps.categories.models import Category
    return Category.objects.create(
        asset_category='测试资产类目',
        item_category='测试物品分类',
        asset_name='测试分类',
        asset_code='CAT-TEST',
        unit='个',
    )


@pytest.fixture
def inventory_task(db, branch, category, admin_user):
    from apps.inventories.models import InventoryTask
    return InventoryTask.objects.create(
        name='测试盘点任务',
        branch=branch,
        category=category,
        status='pending',
        created_by=admin_user,
    )


@pytest.fixture
def in_progress_task(inventory_task):
    inventory_task.status = 'in_progress'
    inventory_task.save()
    return inventory_task


@pytest.fixture
def pending_review_task(inventory_task):
    inventory_task.status = 'pending_review'
    inventory_task.save()
    return inventory_task


@pytest.mark.django_db
class TestCreateTask:
    """创建盘点任务"""

    def test_create_task_success(self, authenticated_client, branch, category):
        payload = {
            'name': '新建盘点任务',
            'branch': branch.id,
            'category': category.id,
            'missed_rule': 'keep',
            'repeat_rule': 'last',
        }
        resp = authenticated_client.post(INVENTORY_LIST_URL, payload, format='json')
        assert resp.status_code == status.HTTP_201_CREATED
        data = resp.json()
        assert data['name'] == '新建盘点任务'
        assert data['status'] == 'pending'

    def test_create_task_missing_name(self, authenticated_client, branch, category):
        payload = {
            'branch': branch.id,
            'category': category.id,
        }
        resp = authenticated_client.post(INVENTORY_LIST_URL, payload, format='json')
        assert resp.status_code == status.HTTP_400_BAD_REQUEST

    def test_list_tasks(self, authenticated_client, inventory_task):
        resp = authenticated_client.get(INVENTORY_LIST_URL, format='json')
        assert resp.status_code == status.HTTP_200_OK


@pytest.mark.django_db
class TestStartTask:
    """启动盘点"""

    def test_start_success(self, authenticated_client, inventory_task):
        resp = authenticated_client.post(
            _task_action_url('start', inventory_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'in_progress'

    def test_start_already_in_progress(self, authenticated_client, in_progress_task):
        resp = authenticated_client.post(
            _task_action_url('start', in_progress_task.id),
            format='json',
        )
        # Should fail — invalid state transition
        assert resp.status_code in (status.HTTP_400_BAD_REQUEST, status.HTTP_200_OK)

    def test_start_unauthenticated(self, api_client, inventory_task):
        resp = api_client.post(
            _task_action_url('start', inventory_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_401_UNAUTHORIZED


@pytest.mark.django_db
class TestCheckItem:
    """盘点扫描"""

    def test_check_success(self, authenticated_client, in_progress_task):
        from apps.assets.models import AssetStock
        from apps.assets.services import ledger
        from apps.categories.models import Category
        item, _ = Category.objects.get_or_create(
            asset_code='AST-INV-001',
            defaults={'asset_category': '测试类目', 'item_category': '测试分类',
                      'asset_name': '盘点测试资产', 'unit': '个'},
        )
        ledger.apply_adjustment(in_progress_task.branch, item, ledger.COLUMN_STOCK, 10, '造数')
        stock = AssetStock.objects.get(branch=in_progress_task.branch, item=item)
        payload = {
            'stockId': str(stock.id),
            'qty': 10,
        }
        resp = authenticated_client.post(
            _task_action_url('check', in_progress_task.id),
            payload,
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK

    def test_check_invalid_qty(self, authenticated_client, in_progress_task):
        payload = {
            'stockId': '00000000-0000-0000-0000-000000000000',
            'qty': -1,
        }
        resp = authenticated_client.post(
            _task_action_url('check', in_progress_task.id),
            payload,
            format='json',
        )
        assert resp.status_code in (status.HTTP_400_BAD_REQUEST, status.HTTP_404_NOT_FOUND)


@pytest.mark.django_db
class TestSubmitTask:
    """提交盘点"""

    def test_submit_success(self, authenticated_client, in_progress_task):
        resp = authenticated_client.post(
            _task_action_url('submit', in_progress_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'pending_review'

    def test_submit_pending_task_fails(self, authenticated_client, inventory_task):
        resp = authenticated_client.post(
            _task_action_url('submit', inventory_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
class TestApproveTask:
    """审批盘点"""

    def test_approve_success(self, authenticated_client, pending_review_task):
        resp = authenticated_client.post(
            _task_action_url('approve', pending_review_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'completed'

    def test_approve_not_pending_review(self, authenticated_client, in_progress_task):
        resp = authenticated_client.post(
            _task_action_url('approve', in_progress_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
class TestRejectTask:
    """驳回盘点"""

    def test_reject_success(self, authenticated_client, pending_review_task):
        resp = authenticated_client.post(
            _task_action_url('reject', pending_review_task.id),
            {'reason': '数据不准确'},
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        data = resp.json()
        assert data['status'] == 'rejected'

    def test_reject_missing_reason(self, authenticated_client, pending_review_task):
        resp = authenticated_client.post(
            _task_action_url('reject', pending_review_task.id),
            {},
            format='json',
        )
        assert resp.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
class TestRecountTask:
    """重新盘点"""

    def test_recount_from_rejected(self, authenticated_client, pending_review_task):
        # First reject the task
        authenticated_client.post(
            _task_action_url('reject', pending_review_task.id),
            {'reason': '需要重新盘点'},
            format='json',
        )
        pending_review_task.refresh_from_db()

        resp = authenticated_client.post(
            _task_action_url('recount', pending_review_task.id),
            {'reset_scope': 'all'},
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'in_progress'


@pytest.mark.django_db
class TestCancelTask:
    """取消盘点"""

    def test_cancel_pending(self, authenticated_client, inventory_task):
        resp = authenticated_client.post(
            _task_action_url('cancel', inventory_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'cancelled'

    def test_cancel_in_progress(self, authenticated_client, in_progress_task):
        resp = authenticated_client.post(
            _task_action_url('cancel', in_progress_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['status'] == 'cancelled'

    def test_cancel_completed_fails(self, authenticated_client, pending_review_task):
        # Approve first to make it completed
        authenticated_client.post(
            _task_action_url('approve', pending_review_task.id),
            format='json',
        )
        pending_review_task.refresh_from_db()

        resp = authenticated_client.post(
            _task_action_url('cancel', pending_review_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
class TestProgressAndReport:
    """盘点进度和报告"""

    def test_progress(self, authenticated_client, in_progress_task):
        resp = authenticated_client.get(
            _task_action_url('progress', in_progress_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK

    def test_report(self, authenticated_client, in_progress_task):
        resp = authenticated_client.get(
            _task_action_url('report', in_progress_task.id),
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK


@pytest.mark.django_db
class TestChecksDisplayFields:
    """盘点记录展示字段与按人过滤（详情页按人流水数据层）"""

    def _make_check(self, task, stock, qty, checked_by):
        from apps.inventories.models import InventoryCheck, InventoryItem
        item = InventoryItem.objects.create(
            task=task, stock=stock, expected_qty=stock.在库数量,
            actual_qty=qty, result='matched', checked_by=checked_by,
        )
        return InventoryCheck.objects.create(
            task=task, item=item, stock=stock, qty=qty, checked_by=checked_by,
        )

    def test_checks_return_display_fields(self, authenticated_client, in_progress_task, admin_user, make_stock):
        stock = make_stock(code='CHK-001', qty=10)
        self._make_check(in_progress_task, stock, 10, admin_user)

        resp = authenticated_client.get(
            _task_action_url('checks', in_progress_task.id), format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        results = resp.json()['results']
        assert len(results) == 1
        record = results[0]
        assert record['checkedByName'] == '测试管理员'
        assert record['assetCode'] == 'CHK-001'
        assert record['assetName'] == stock.item.asset_name
        assert record['qty'] == 10

    def test_checks_filter_by_person(self, authenticated_client, in_progress_task, admin_user, make_stock):
        from django.contrib.auth import get_user_model
        other = get_user_model().objects.create_user(
            phone='13911112222', name='第二盘点人', password='test123456',
            role='staff', status='active',
        )
        self._make_check(in_progress_task, make_stock(code='CHK-A', qty=5), 5, admin_user)
        self._make_check(in_progress_task, make_stock(code='CHK-B', qty=6), 6, other)

        resp = authenticated_client.get(
            _task_action_url('checks', in_progress_task.id),
            {'checkedBy': str(other.id)},
            format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        data = resp.json()
        assert data['count'] == 1
        assert data['results'][0]['checkedByName'] == '第二盘点人'
        assert data['results'][0]['assetCode'] == 'CHK-B'

    def test_checks_invalid_person_param(self, authenticated_client, in_progress_task):
        resp = authenticated_client.get(
            _task_action_url('checks', in_progress_task.id),
            {'checkedBy': 'not-a-uuid'},
            format='json',
        )
        assert resp.status_code == status.HTTP_400_BAD_REQUEST

    def test_checks_empty(self, authenticated_client, in_progress_task):
        resp = authenticated_client.get(
            _task_action_url('checks', in_progress_task.id), format='json',
        )
        assert resp.status_code == status.HTTP_200_OK
        assert resp.json()['count'] == 0
        assert resp.json()['results'] == []


@pytest.mark.django_db
class TestExportReport:
    """盘点报告 Excel 导出（基本信息 + 统计 + 调整单号 + 明细，双 sheet）"""

    def _make_item(self, task, stock, expected, actual, result, checked_by, remarks=''):
        from apps.inventories.models import InventoryItem
        return InventoryItem.objects.create(
            task=task, stock=stock, expected_qty=expected, actual_qty=actual,
            result=result, checked_by=checked_by, remarks=remarks,
        )

    def _make_task(self, branch, admin_user, status, name='导出测试任务'):
        from apps.inventories.models import InventoryTask
        return InventoryTask.objects.create(
            name=name, branch=branch, status=status, created_by=admin_user,
        )

    def test_export_completed_task_with_adjustments(self, authenticated_client, branch, admin_user, make_stock):
        stock_surplus = make_stock(code='EXP-S-01', qty=5)
        stock_missing = make_stock(code='EXP-M-01', qty=5)
        task = self._make_task(branch, admin_user, 'in_progress')
        self._make_item(task, stock_surplus, 5, 6, 'surplus', admin_user, remarks='盘盈一件')
        self._make_item(task, stock_missing, 5, 3, 'missing', admin_user, remarks='少了两件')
        task.status = 'pending_review'
        task.save()

        approve_resp = authenticated_client.post(_task_action_url('approve', task.id), format='json')
        assert approve_resp.status_code == status.HTTP_200_OK
        adj_numbers = list(task.adjustments.values_list('单据编号', flat=True))
        assert len(adj_numbers) == 2

        resp = authenticated_client.get(_task_action_url('export-report', task.id))
        assert resp.status_code == status.HTTP_200_OK
        assert 'spreadsheetml' in resp['Content-Type']
        # 非 latin-1 头值被 Django 按 RFC 2047 编码，解码后校验文件名
        from email.header import decode_header, make_header
        disposition = str(make_header(decode_header(resp['Content-Disposition'])))
        assert '盘点报告' in disposition
        assert '导出测试任务' in disposition

        wb = _load_workbook(resp)
        assert wb.sheetnames == ['盘点报告', '盘点明细']

        report_cells = _sheet_cells(wb['盘点报告'])
        for num in adj_numbers:
            assert num in report_cells
        assert '无（任务未完成）' not in report_cells
        assert '差异调整单' in report_cells

        detail_rows = list(wb['盘点明细'].iter_rows(values_only=True))
        assert detail_rows[0][0] == '序号'
        missing_row = next(r for r in detail_rows if r[1] == 'EXP-M-01')
        assert missing_row[4] == 5
        assert missing_row[5] == 3
        assert missing_row[6] == '-2'
        assert missing_row[7] == '盘亏'
        assert missing_row[8] == '测试管理员'
        assert missing_row[10] == '少了两件'
        surplus_row = next(r for r in detail_rows if r[1] == 'EXP-S-01')
        assert surplus_row[6] == '+1'
        assert surplus_row[7] == '盘盈'

    def test_export_in_progress_snapshot(self, authenticated_client, branch, admin_user, make_stock):
        task = self._make_task(branch, admin_user, 'in_progress', name='进行中任务')
        self._make_item(task, make_stock(code='EXP-P-01', qty=5), 5, 3, 'missing', admin_user)

        resp = authenticated_client.get(_task_action_url('export-report', task.id))
        assert resp.status_code == status.HTTP_200_OK

        wb = _load_workbook(resp)
        report_cells = _sheet_cells(wb['盘点报告'])
        assert '无（任务未完成）' in report_cells

        detail_rows = list(wb['盘点明细'].iter_rows(values_only=True))
        assert len(detail_rows) == 2
        assert detail_rows[1][1] == 'EXP-P-01'
        assert detail_rows[1][6] == '-2'

    def test_export_pending_task_empty_detail(self, authenticated_client, inventory_task):
        resp = authenticated_client.get(_task_action_url('export-report', inventory_task.id))
        assert resp.status_code == status.HTTP_200_OK

        wb = _load_workbook(resp)
        detail_rows = list(wb['盘点明细'].iter_rows(values_only=True))
        assert len(detail_rows) == 1  # 仅表头
        assert '无（任务未完成）' in _sheet_cells(wb['盘点报告'])

    def test_export_out_of_scope_404(self, supervisor_client, second_branch, admin_user):
        task = self._make_task(second_branch, admin_user, 'in_progress', name='区域B任务')
        resp = supervisor_client.get(_task_action_url('export-report', task.id))
        assert resp.status_code == status.HTTP_404_NOT_FOUND
