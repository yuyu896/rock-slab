"""
Comprehensive tests for all import/export/template endpoints across 7 modules.
Covers: Asset, FixedAsset, Category, Transfer (4 types), Inventory.
"""
import io
import pytest
from datetime import date

import openpyxl
from rest_framework import status

from conftest import _client_for


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(headers, rows=None):
    """Build an in-memory Excel file with given headers and rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in (rows or []):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'test.xlsx'
    return buf


def _parse_excel_response(response):
    """Parse an Excel response into (headers, rows)."""
    assert response.status_code == 200, f"Expected 200, got {response.status_code}"
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    headers = list(data[0]) if data else []
    rows = [list(r) for r in data[1:]] if len(data) > 1 else []
    return headers, rows


def _upload_url(client, url, buf, params=None):
    """Upload an Excel buffer to the given URL."""
    return client.post(url, {'file': buf}, format='multipart', QUERY_STRING=params or '')


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def admin_client(admin_user):
    return _client_for(admin_user)


@pytest.fixture
def test_branch(db):
    from apps.organizations.models import Region, Team, Branch
    region = Region.objects.create(name='测试区域', code='TEST', status='active')
    team = Team.objects.create(name='测试行政组', region=region, status='active')
    branch = Branch.objects.create(
        name='测试分公司', code='TB001', team=team,
        address='测试地址', phone='010-12345678',
    )
    return branch


@pytest.fixture
def test_category(db):
    from apps.categories.models import Category
    return Category.objects.create(
        asset_category='测试类目', item_category='测试分类',
        asset_name='测试资产', asset_code='TC001', unit='台',
    )


@pytest.fixture
def test_asset(db, test_branch):
    from apps.assets.models import Asset
    return Asset.objects.create(
        序号=1, 分公司=test_branch.name, 分公司编号=test_branch.code,
        branch=test_branch, 资产编号='TA001', 资产类目='测试类目',
        物品分类='测试分类', 资产名称='测试资产', 数量=5, 当前状态='在库',
    )


@pytest.fixture
def ta001_category(db):
    from apps.categories.models import Category
    return Category.objects.create(
        asset_category='测试类目', item_category='测试分类',
        asset_name='测试资产', asset_code='TA001', unit='台',
    )


# ===========================================================================
# 2. Asset module
# ===========================================================================

ASSET_HEADERS = [
    '分公司', '资产编号', '分公司编号', '资产类目',
    '电脑序列号', '供应商', '物品分类', '资产名称', '图片',
    '入库日期', '是否租用', '数量', '规格', '单价',
    '购入金额', '出库日期', '所属部门', '使用人', '当前状态',
    '是否充足', '备注',
]


class TestAssetTemplate:
    def test_download_template(self, admin_client):
        resp = admin_client.get('/api/assets/template')
        headers, rows = _parse_excel_response(resp)
        assert rows == []
        expected = ['分公司', '资产编号', '资产类目', '物品分类', '资产名称', '入库日期', '是否租用',
                    '数量', '规格', '单价', '购入金额', '出库日期', '所属部门', '当前状态', '备注']
        for h in expected:
            assert h in headers, f"Missing header: {h}"
        assert len(headers) == len(expected)
        assert '供应商' not in headers


class TestAssetImport:
    """资产明细导入已随 Asset 冻结下线（410），改走台账增量导入。"""

    def test_import_returns_410(self, admin_client):
        buf = _make_xlsx(ASSET_HEADERS, [])
        resp = _upload_url(admin_client, '/api/assets/import', buf)
        assert resp.status_code == status.HTTP_410_GONE



class TestAssetExport:
    def test_export_with_data(self, admin_client, test_asset):
        resp = admin_client.get('/api/assets/export')
        headers, rows = _parse_excel_response(resp)
        assert len(rows) >= 1
        assert test_asset.资产编号 in [r[2] for r in rows]

    def test_export_empty(self, admin_client, db):
        from apps.assets.models import Asset
        Asset.objects.all().delete()
        resp = admin_client.get('/api/assets/export')
        headers, rows = _parse_excel_response(resp)
        assert len(rows) == 0


# ===========================================================================
# 3. FixedAsset module
# ===========================================================================

class TestFixedAssetEndpointsFrozen:
    """P2 第二刀：实例导入/模板下线——出生=采购单，存量=迁移。"""

    def test_template_endpoint_gone(self, admin_client):
        resp = admin_client.get('/api/assets/fixed-assets/template')
        assert resp.status_code in (404, 410)

    def test_import_endpoint_gone(self, admin_client):
        buf = _make_xlsx(['资产编号', '名称'], [['TA001', 'x']])
        resp = _upload_url(admin_client, '/api/assets/fixed-assets/import', buf)
        assert resp.status_code == status.HTTP_410_GONE
        assert '采购入库单' in resp.data['detail']


# ===========================================================================
# 4. Category module
# ===========================================================================

CATEGORY_TEMPLATE_HEADERS = ['资产类目', '物品分类', '资产名称', '资产编号', '计量单位', '警戒线', '备注']
CATEGORY_EXPORT_HEADERS = ['资产类目', '物品分类', '资产名称', '资产编号', '计量单位', '资产数量', '在库数量', '警戒线', '备注']


class TestCategoryTemplate:
    def test_download_template(self, admin_client):
        resp = admin_client.get('/api/categories/template')
        headers, rows = _parse_excel_response(resp)
        assert rows == []
        for h in CATEGORY_TEMPLATE_HEADERS:
            assert h in headers, f"Missing header: {h}"


class TestCategoryImport:
    def test_import_valid_data(self, admin_client):
        rows = [
            ['办公类', '电子设备', '笔记本电脑', 'CAT-L001', '台', 10, '办公用'],
            ['办公类', '家具', '办公桌', 'CAT-D001', '张', '', ''],
        ]
        buf = _make_xlsx(CATEGORY_TEMPLATE_HEADERS, rows)
        resp = _upload_url(admin_client, '/api/categories/import', buf)
        assert resp.status_code == status.HTTP_200_OK
        assert resp.data['imported'] == 2
        from apps.categories.models import Category
        c = Category.objects.get(asset_code='CAT-L001')
        assert c.asset_name == '笔记本电脑'
        assert c.unit == '台'

    def test_import_duplicate_asset_code_updates(self, admin_client, test_category):
        """Category import uses update_or_create, so duplicate asset_code should update."""
        rows = [
            ['更新类目', '更新分类', '更新资产', test_category.asset_code, '套', '', ''],
        ]
        buf = _make_xlsx(CATEGORY_TEMPLATE_HEADERS, rows)
        resp = _upload_url(admin_client, '/api/categories/import', buf)
        assert resp.status_code == status.HTTP_200_OK
        assert resp.data['imported'] == 1
        test_category.refresh_from_db()
        assert test_category.asset_name == '更新资产'
        assert test_category.unit == '套'


class TestCategoryExport:
    def test_export_with_data(self, admin_client, test_category):
        resp = admin_client.get('/api/categories/export')
        headers, rows = _parse_excel_response(resp)
        assert len(rows) >= 1


# ===========================================================================
# 5-8. Transfer module (purchase/assign/transfer/recovery)
# ===========================================================================

# Template headers must match the TYPE_TEMPLATES in TransferViewSet exactly
# P2 明细行化后断言分三层：check_fields=单头字段；line_check_fields=明细行字段；
# item_check_fields=字典联查回显字段（单位/类目等行内不存，取自品目字典）。
TRANSFER_TYPE_TEMPLATES = {
    'purchase': {
        'template_headers': ['采购日期', '分公司', '资产编号', '物品名称', '规格型号', '图片',
                             '供应商', '采购数量', '单价', '总金额', '需求部门', '采购经办人', '备注'],
        'sample_row': ['2026-03-01', '测试分公司', 'PUR-001', '采购物品A', '规格X', '',
                       '供应商A', 10, 50.0, 500.0, '研发部', '李四', '采购备注'],
        'check_fields': {'供应商': '供应商A', '需求部门': '研发部', '采购经办人': '李四'},
        'line_check_fields': {},
        'item_check_fields': {},
    },
    'assign': {
        'template_headers': ['分公司', '日期', '资产编号', '领用物品', '领用数量', '用途', '领用部门', '备注'],
        'sample_row': ['测试分公司', '2026-03-01', 'AST-TEST-001', '领用物品B', 5, '办公用', '行政部', ''],
        'check_fields': {'用途': '办公用'},
        'line_check_fields': {'数量': 5},
        'item_check_fields': {},
    },
    'transfer': {
        'template_headers': ['调拨日期', '调出分公司', '调出部门', '调入分公司', '调入部门',
                             '资产编号', '资产名称', '规格型号', '调拨数量', '调拨原因',
                             '调出负责人', '调入负责人', '备注'],
        'sample_row': ['2026-03-01', '上海分公司', '行政部', '杭州分公司', '研发部',
                       'TRF-001', '调拨物品C', '规格Y', 3, '部门调整', '王五', '赵六', ''],
        'check_fields': {'调出分公司': '上海分公司'},
        'line_check_fields': {'数量': 3},
        'item_check_fields': {},
    },
    'recovery': {
        # Must match TYPE_TEMPLATES['recovery']['headers'] in views.py exactly
        'template_headers': ['分公司', '资产编号', '资产类目', '物品分类', '资产名称', '回收分类',
                             '入库日期', '数量', '单位', '规格', '出库日期', '所属部门',
                             '存放位置', '经办人', '备注'],
        # Must match import parsing: row[0]=分公司, row[1]=资产编号, row[2]=资产类目,
        # row[3]=物品分类, row[4]=资产名称, row[5]=回收分类, row[6]=入库日期(→调拨日期),
        # row[7]=数量, row[8]=单位, row[9]=规格, row[10]=出库日期, row[11]=所属部门,
        # row[12]=?(unused), row[13]=存放位置, row[14]=经办人(→采购经办人), row[15]=备注
        'sample_row': ['测试分公司', 'REC-001', '电子设备', '电脑', '回收电脑',
                       '闲置回收', '2026-03-01', 2, '台', '型号Z', '2026-03-05',
                       '行政部', '', '仓库B', '张采购', '回收备注'],
        'check_fields': {'回收分类': '闲置回收', '采购经办人': '张采购'},
        'line_check_fields': {'存放位置': '仓库B'},
        'item_check_fields': {'unit': '台', 'asset_category': '电子设备', 'item_category': '电脑'},
    },
}


class TestTransferTemplates:
    @pytest.mark.parametrize('ttype', ['purchase', 'assign', 'transfer', 'recovery'])
    def test_download_template(self, admin_client, ttype):
        tpl = TRANSFER_TYPE_TEMPLATES[ttype]
        resp = admin_client.get(f'/api/transfers/template?type={ttype}')
        headers, rows = _parse_excel_response(resp)
        assert rows == []
        for h in tpl['template_headers']:
            assert h in headers, f"[{ttype}] Missing header: {h}"


def _seed_recovery_dictionary():
    """预置 REC-001 字典展示值（单位/类目），与导入模板样例行一致——P2 起这些值取自字典而非行内。"""
    from apps.categories.models import Category
    Category.objects.filter(asset_code='REC-001').update(
        asset_category='电子设备', item_category='电脑', unit='台',
    )


class TestTransferImport:
    @pytest.mark.parametrize('ttype', ['purchase', 'assign', 'transfer', 'recovery'])
    def test_import_valid_data(self, admin_client, ttype, test_branch):
        from apps.transfers.models import Transfer
        from apps.organizations.models import Branch
        # transfer 类型引用 上海/杭州分公司，确保其存在于组织架构（导入现校验分公司存在性）
        Branch.objects.create(name='上海分公司', code='SH001', team=test_branch.team)
        Branch.objects.create(name='杭州分公司', code='HZ001', team=test_branch.team)
        if ttype == 'recovery':
            _seed_recovery_dictionary()
        tpl = TRANSFER_TYPE_TEMPLATES[ttype]
        buf = _make_xlsx(tpl['template_headers'], [tpl['sample_row']])
        resp = _upload_url(admin_client, '/api/transfers/import', buf, params=f'type={ttype}')
        assert resp.status_code == status.HTTP_200_OK
        assert resp.data['imported'] >= 1
        t = Transfer.objects.filter(action_type=ttype).last()
        assert t is not None, f"No transfer found for type {ttype}"
        for field, expected in tpl['check_fields'].items():
            actual = getattr(t, field)
            if isinstance(expected, (int, float)):
                assert float(actual) == float(expected), f"[{ttype}] {field}: {actual} != {expected}"
            else:
                assert actual == expected, f"[{ttype}] {field}: '{actual}' != '{expected}'"
        line = t.lines.select_related('item').first()
        assert line is not None, f"[{ttype}] No transfer line found"
        for field, expected in tpl['line_check_fields'].items():
            actual = getattr(line, field)
            if isinstance(expected, (int, float)):
                assert float(actual) == float(expected), f"[{ttype}] line {field}: {actual} != {expected}"
            else:
                assert actual == expected, f"[{ttype}] line {field}: '{actual}' != '{expected}'"
        for field, expected in tpl['item_check_fields'].items():
            actual = getattr(line.item, field)
            assert actual == expected, f"[{ttype}] item {field}: '{actual}' != '{expected}'"

    def test_recovery_import_all_fields(self, admin_client, test_branch):
        """Verify all recovery-specific fields survive import (单头 + 明细行 + 字典联查)."""
        from apps.transfers.models import Transfer
        _seed_recovery_dictionary()
        tpl = TRANSFER_TYPE_TEMPLATES['recovery']
        buf = _make_xlsx(tpl['template_headers'], [tpl['sample_row']])
        resp = _upload_url(admin_client, '/api/transfers/import', buf, params='type=recovery')
        assert resp.status_code == status.HTTP_200_OK
        t = Transfer.objects.filter(action_type='recovery').last()
        assert t is not None
        assert t.回收分类 == '闲置回收'
        assert t.采购经办人 == '张采购'
        line = t.lines.select_related('item').first()
        assert line is not None
        assert line.存放位置 == '仓库B'
        assert line.item.unit == '台'
        assert line.item.asset_category == '电子设备'
        assert line.item.item_category == '电脑'


    def test_import_rejects_unknown_branch(self, admin_client):
        # transfer 类型：调出分公司填写不存在的名称 → 该行被拒
        tpl = TRANSFER_TYPE_TEMPLATES['transfer']
        row = list(tpl['sample_row'])
        row[1] = '不存在的分公司'  # 调出分公司
        buf = _make_xlsx(tpl['template_headers'], [row])
        resp = _upload_url(admin_client, '/api/transfers/import', buf, params='type=transfer')
        assert resp.status_code == status.HTTP_200_OK
        assert resp.data['imported'] == 0
        assert any('不存在的分公司' in e for e in resp.data['errors'])

    def test_import_rejects_empty_branch(self, admin_client):
        # transfer 类型：调出分公司为空 → 该行被拒
        tpl = TRANSFER_TYPE_TEMPLATES['transfer']
        row = list(tpl['sample_row'])
        row[1] = ''  # 调出分公司为空
        buf = _make_xlsx(tpl['template_headers'], [row])
        resp = _upload_url(admin_client, '/api/transfers/import', buf, params='type=transfer')
        assert resp.status_code == status.HTTP_200_OK
        assert resp.data['imported'] == 0
        assert any('调出分公司为空' in e for e in resp.data['errors'])


class TestTransferExport:
    @pytest.mark.parametrize('ttype', ['purchase', 'assign', 'transfer', 'recovery'])
    def test_export_filters_by_type(self, admin_client, ttype):
        from apps.transfers.models import Transfer, TransferLine
        from apps.categories.models import Category
        # Create one record of each type（单头 + 一条明细行，导出按行展开）
        for tt in ['purchase', 'assign', 'transfer', 'recovery']:
            item = Category.objects.create(
                asset_category='测试类目', item_category='测试分类',
                asset_name=f'导出测试-{tt}', asset_code=f'EXP-{tt}-001', unit='个',
            )
            t = Transfer.objects.create(
                调拨日期=date(2026, 3, 1),
                action_type=tt,
            )
            TransferLine.objects.create(transfer=t, item=item, 行号=1, 数量=1)
        resp = admin_client.get(f'/api/transfers/export?type={ttype}')
        headers, rows = _parse_excel_response(resp)
        assert len(rows) >= 1


# ===========================================================================
# 9. Inventory module
# ===========================================================================

class TestInventoryImportExport:
    def test_download_template(self, admin_client, test_asset, admin_user):
        from apps.inventories.models import InventoryTask
        task = InventoryTask.objects.create(
            name='导入测试盘点',
            status='in_progress',
            created_by=admin_user,
        )
        resp = admin_client.get(f'/api/inventories/{task.id}/import-template')
        assert resp.status_code == status.HTTP_200_OK
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.active is not None

    def test_import_results(self, admin_client, test_asset, admin_user):
        from apps.inventories.models import InventoryTask, InventoryItem
        task = InventoryTask.objects.create(
            name='结果导入测试', status='in_progress', created_by=admin_user,
        )
        InventoryItem.objects.create(
            task=task, asset=test_asset,
            expected_qty=5, actual_qty=None, result='unchecked', check_count=0,
        )
        # Inventory import expects specific columns: match the actual template
        # row[1]=资产编号, row[5]=实盘数量
        headers = ['序号', '资产编号', '资产名称', '规格', '账面数量', '实盘数量', '备注']
        rows = [[1, test_asset.资产编号, test_asset.资产名称, '', 5, 5, '盘点正常']]
        buf = _make_xlsx(headers, rows)
        resp = _upload_url(admin_client, f'/api/inventories/{task.id}/import-result', buf)
        assert resp.status_code == status.HTTP_200_OK
        item = InventoryItem.objects.get(task=task, asset=test_asset)
        assert item.actual_qty == 5
        assert item.result == 'matched'


# ===========================================================================
# 10. Edge cases
# ===========================================================================

class TestImportEdgeCases:
    """资产导入下线后，边界校验仅剩 410 语义；台账/固定资产导入边界见各自测试。"""

    def test_asset_import_wrong_extension_returns_410(self, admin_client):
        from io import BytesIO
        buf = BytesIO(b'not excel')
        buf.name = 'test.txt'
        resp = admin_client.post('/api/assets/import', {'file': buf}, format='multipart')
        assert resp.status_code == status.HTTP_410_GONE


